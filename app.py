"""
Open SEO Crawler — a fast, concurrent SEO-focused web crawler.

Single-file Flask app. Run with:
    python3 app.py
Then open http://localhost:5002/

Features:
  - Concurrent threaded crawl with per-host politeness
  - CMS detection (Shopify, WordPress + Yoast/Rank Math, Webflow, Wix, Squarespace, etc.)
  - Retry on 429 / 5xx / connection errors with exponential backoff
  - Duplicate title / meta / H1 / body detection
  - Orphan page detection (sitemap vs crawled URLs)
  - Redirect chain detection
  - Severity-tagged issues (error / warning / info) with source citations
  - Optional JS rendering via Playwright (install separately)
"""
from flask import Flask, render_template, request, Response, stream_with_context, jsonify, send_file
import requests
import json
import os
import re as _re
import time
import logging
import threading
from collections import deque, defaultdict as _dd
from urllib.parse import urlparse, urljoin, urlunparse, parse_qs, urlencode
from bs4 import BeautifulSoup

app = Flask(__name__, static_folder='static', template_folder='templates')
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(name)s] %(levelname)s: %(message)s')

# crawl_id -> {'include': [patterns], 'exclude': [patterns]}. Workers read on
# every URL so the Apply button can mutate rules mid-crawl. Replace list refs
# atomically (do not mutate in place) so concurrent readers see a consistent
# snapshot. Cleared when the crawl ends.
ACTIVE_CRAWL_RULES = {}

# crawl_id -> {'max_pages': int, 'continue_event': Event, 'finalize': bool}.
# Lets /crawl/continue bump the page cap (or finalize) when the crawler hits
# the limit with URLs still queued. The generator waits on continue_event.
ACTIVE_CRAWL_LIMITS = {}

# crawl_id -> snapshot saved when the SSE generator hits GeneratorExit
# (network drop, browser close, manual stop). Lets the user pick up from where
# the crawl left off via /crawl with resume_crawl_id set. Auto-prunes anything
# older than SUSPENDED_CRAWL_TTL on access.
SUSPENDED_CRAWLS = {}
SUSPENDED_CRAWL_TTL = 30 * 60  # 30 minutes


def _http_get(url, **kwargs):
    """requests.get that transparently retries with verify=False on an SSL
    cert-chain failure (incomplete chain / untrusted or self-signed cert).
    Lots of sites load fine in browsers but serve a broken chain; without this,
    robots.txt and sitemap discovery silently fail and the user sees
    "no sitemap found". The returned Response carries .ssl_bypassed."""
    try:
        r = requests.get(url, **kwargs)
        r.ssl_bypassed = False
        return r
    except requests.exceptions.SSLError:
        try:
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        except Exception:
            pass
        kwargs['verify'] = False
        r = requests.get(url, **kwargs)
        r.ssl_bypassed = True
        return r


def _http_head(url, **kwargs):
    """HEAD counterpart to _http_get with the same SSL-fallback behaviour."""
    try:
        r = requests.head(url, **kwargs)
        r.ssl_bypassed = False
        return r
    except requests.exceptions.SSLError:
        try:
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        except Exception:
            pass
        kwargs['verify'] = False
        r = requests.head(url, **kwargs)
        r.ssl_bypassed = True
        return r


def _robots_pattern_match(pattern, url):
    """robots.txt-style path pattern matcher (Google's spec).

    Wildcards: ``*`` matches any sequence, ``$`` at end anchors end of URL.
    Everything else is literal — including ``?``. Match is anchored at the
    start of the path; a leading ``*`` lets it match anywhere. Tests path+query
    first, then full URL so users can paste either form.
    """
    if not pattern:
        return False
    p = pattern.strip()
    if not p:
        return False
    end_anchor = p.endswith('$')
    if end_anchor:
        p = p[:-1]
    rx = '.*'.join(_re.escape(part) for part in p.split('*'))
    if end_anchor:
        rx += r'\Z'
    rx = '^' + rx
    try:
        prog = _re.compile(rx)
    except _re.error:
        return False
    parsed = urlparse(url)
    path_q = parsed.path + (('?' + parsed.query) if parsed.query else '')
    if prog.match(path_q):
        return True
    if prog.match(url):
        return True
    return False


def _build_robots_checker(robots_text):
    """Parse robots.txt and return a fn(url) -> bool that honours
    Google's wildcard spec (* and $) and longest-match-wins precedence.

    Python's stdlib urllib.robotparser does NOT support wildcards or end-
    of-URL anchors, so rules like 'Disallow: *?swoof*' or 'Disallow: *?*'
    silently turn into no-ops and the crawler picks up the dynamic URLs
    they were meant to keep out. Common on WooCommerce sites with the
    WOOF/Yoast filter combo. We use the existing _robots_pattern_match
    (which is already Google-spec) and apply longest-pattern-wins so a
    more specific Allow can lift a broader Disallow.

    Reads the User-agent: * block only - we crawl as a generic bot.
    """
    if not robots_text:
        return lambda u: True
    allows, disallows = [], []
    in_star_block = False
    for line in robots_text.splitlines():
        line = line.split('#', 1)[0].strip()
        if not line or ':' not in line:
            continue
        key, _, value = line.partition(':')
        key = key.strip().lower()
        value = value.strip()
        if key == 'user-agent':
            in_star_block = (value == '*')
            continue
        if not in_star_block:
            continue
        if key == 'disallow' and value:
            disallows.append(value)
        elif key == 'allow' and value:
            allows.append(value)
    if not disallows and not allows:
        return lambda u: True

    def can_fetch(url):
        best_allow = -1
        best_disallow = -1
        for a in allows:
            if len(a) > best_allow and _robots_pattern_match(a, url):
                best_allow = len(a)
        for d in disallows:
            if len(d) > best_disallow and _robots_pattern_match(d, url):
                best_disallow = len(d)
        # Tie or longer Allow wins; only block when Disallow is strictly
        # more specific. This matches Google's robots.txt parser.
        return best_disallow <= best_allow

    return can_fetch


# Well-known AI crawler user-agents (2026). Blocking these removes the site from
# AI answer engines (ChatGPT, Claude, Perplexity, Google AI Overviews) and AI
# training sets. The search/citation bots (OAI-SearchBot, Claude-SearchBot,
# PerplexityBot, ChatGPT-User) are the ones that actually feed live AI answers;
# the rest are training crawlers. Grouped only for readability — detection is flat.
_AI_CRAWLER_UAS = [
    'GPTBot', 'OAI-SearchBot', 'ChatGPT-User',            # OpenAI
    'ClaudeBot', 'Claude-Web', 'Claude-SearchBot',         # Anthropic
    'Claude-User', 'anthropic-ai',
    'Google-Extended',                                     # Google AI / Gemini
    'PerplexityBot', 'Perplexity-User',                    # Perplexity
    'CCBot',                                               # Common Crawl (feeds many models)
    'Bytespider',                                          # ByteDance
    'Applebot-Extended',                                   # Apple Intelligence
    'Amazonbot',                                           # Amazon
    'Meta-ExternalAgent', 'meta-externalagent',            # Meta AI
    'cohere-ai', 'Diffbot', 'AI2Bot', 'MistralAI-User',
    'DuckAssistBot', 'YouBot', 'Timpibot', 'ImagesiftBot', 'Omgilibot',
]
# Classic search-engine crawlers. Blocking these is almost always a mistake and
# removes the site from normal (blue-link) search results entirely.
_SEARCH_ENGINE_UAS = ['Googlebot', 'Bingbot', 'Slurp', 'DuckDuckBot', 'Baiduspider', 'YandexBot']


def _parse_robots_groups(robots_text):
    """Parse robots.txt into {lower_user_agent: {'disallow': [...], 'allow': [...]}}.

    Consecutive `User-agent:` lines share the rule block that follows them, per the
    robots.txt spec. Unlike _build_robots_checker (which only reads the `*` group to
    drive crawling), this keeps every named group so we can audit per-bot blocking.
    """
    groups = {}
    current = []
    starting_group = True
    for raw in (robots_text or '').splitlines():
        line = raw.split('#', 1)[0].strip()
        if not line or ':' not in line:
            continue
        key, _, value = line.partition(':')
        key = key.strip().lower()
        value = value.strip()
        if key == 'user-agent':
            if not starting_group:
                current = []
                starting_group = True
            ua = value.lower()
            current.append(ua)
            groups.setdefault(ua, {'disallow': [], 'allow': []})
        elif key in ('disallow', 'allow'):
            starting_group = False
            for ua in current:
                groups[ua][key].append(value)
        else:
            starting_group = False
    return groups


def _robots_root_blocked(group):
    """True if the group disallows the site root ('/') with no Allow: / lifting it."""
    if not group:
        return False
    blocked = any(d.strip() in ('/', '/*') for d in group.get('disallow', []))
    if blocked and any(a.strip() == '/' for a in group.get('allow', [])):
        return False
    return blocked


def _analyze_robots_txt(robots_text):
    """Inspect robots.txt for crawler-blocking that hurts SEO / AI visibility.

    Returns a list of human-readable issue strings. Both blocking AI crawlers and
    blocking search engines are surfaced as red errors (see the sev() classifier);
    AI blocking is the bigger commercial problem as AI answer engines grow.
    """
    issues = []
    if not robots_text:
        return issues
    groups = _parse_robots_groups(robots_text)
    star = groups.get('*')

    def is_blocked(ua):
        g = groups.get(ua.lower())
        if g is not None:
            return _robots_root_blocked(g)
        return _robots_root_blocked(star)  # no own group -> falls back to the * group

    # Cloudflare-style Content-Signal opt-out (e.g. "search=yes,ai-train=no").
    signal_block = False
    for raw in robots_text.splitlines():
        l = raw.split('#', 1)[0].strip().lower().replace(' ', '')
        if l.startswith('content-signal') and ('ai-train=no' in l or 'ai-input=no' in l):
            signal_block = True

    # AI crawlers (de-duped, original casing preserved).
    seen, ai_names = set(), []
    for ua in _AI_CRAWLER_UAS:
        if is_blocked(ua) and ua.lower() not in seen:
            seen.add(ua.lower())
            ai_names.append(ua)
    if ai_names or signal_block:
        if ai_names:
            shown = ai_names[:10]
            if len(ai_names) > 10:
                shown.append(f'+{len(ai_names) - 10} more')
            tail = ' (+ Content-Signal ai-train=no)' if signal_block else ''
            issues.append('AI crawlers blocked in robots.txt — ' + ', '.join(shown) + tail)
        else:
            issues.append('AI crawlers blocked in robots.txt — Content-Signal set to ai-train/ai-input "no"')

    # Classic search engines.
    blocked_se = [ua for ua in _SEARCH_ENGINE_UAS if is_blocked(ua)]
    if blocked_se:
        issues.append('Search engines blocked in robots.txt — ' + ', '.join(blocked_se))
    elif _robots_root_blocked(star):
        issues.append('Search engines blocked in robots.txt — User-agent: * Disallow: /')

    return issues


_DUP_NOISE_PARAMS = frozenset({
    'add-to-cart', 'replytocom', 'fbclid', 'gclid', 'gad_source', 'gbraid',
    'wbraid', 'mc_cid', 'mc_eid', '_ga', 'msclkid', 'yclid', 'dclid',
    'igshid', 'srsltid', 'ref', 'ref_src', 'ref_url',
    # WooCommerce SWOOF / product filter plugins — ?swoof=1&pa_cube-size=28mm
    # &product_cat=drawer&really_curr_tax=63-product_cat etc. Each combination
    # is a separate URL but renders the same template; collapsing them stops
    # filter permutations from polluting duplicate-meta/title groups.
    'swoof', 'really_curr_tax', 'product_cat', 'product_tag',
    'orderby', 'min_price', 'max_price', 'rating_filter', 'filter_size',
    'filter_color',
})


def _normalize_url_for_dup(url):
    """Dedup key inside `_group_by`. Strips the entire query string + collapses
    pagination tails so filter/sort/search variants of the same page (e.g.
    /faq/ vs /faq/?category=planning) collapse to one entry. The grouper has
    already bucketed by identical meta/title/H1/body; URLs with truly distinct
    content sit in different buckets so this can't cause false collapses."""
    if not url:
        return url
    try:
        parsed = urlparse(url)
    except Exception:
        return url
    # Collapse http vs https — the same URL on different schemes is the same
    # page, not a duplicate-title issue. (HTTP-only pages are surfaced
    # separately by the security report.)
    scheme = 'https'
    netloc = (parsed.netloc or '').lower()
    if netloc.startswith('www.'):
        netloc = netloc[4:]
    path = parsed.path or '/'
    path = _re.sub(r'/page/\d+/?$', '/', path)
    path = _re.sub(r'/comment-page-\d+/?$', '/', path)
    # Strip trailing whitespace / %20 / slashes in any combo so /foo,
    # /foo/, /foo%20, /foo%20/ all collapse. Common on Shopify when an
    # internal <a href> has a trailing space → encoded as %20.
    while path and path != '/':
        low = path.lower()
        if low.endswith('%20'):
            path = path[:-3]
        elif path[-1].isspace() or path[-1] == '/':
            path = path[:-1]
        else:
            break
    if not path:
        path = '/'
    return urlunparse((scheme, netloc, path, '', '', ''))


STATIC_VERSION = str(int(time.time()))


def _local_commit_sha():
    """Short SHA of the currently-installed build, or 'dev' if not a
    git checkout. Resolved per-call (not cached at startup) so devs
    don't get a stale 'update available' banner the moment they push
    a commit without restarting the dev server."""
    try:
        import subprocess as _sp
        out = _sp.run(['git', '-C', os.path.dirname(os.path.abspath(__file__)),
                       'rev-parse', '--short', 'HEAD'],
                      capture_output=True, text=True, timeout=2)
        if out.returncode == 0 and out.stdout.strip():
            return out.stdout.strip()
    except Exception:
        pass
    return 'dev'


@app.route('/version')
def version():
    """Local build SHA, served to the UI for display + comparison
    against the GitHub master HEAD (client-side fetch)."""
    return jsonify({
        'sha': _local_commit_sha(),
        'repo': 'puneetindersingh/open-seo-crawler',
    })


@app.route('/update', methods=['POST'])
def update_self():
    """Reconcile the local checkout to origin/master and report the result.
    Restart is handled separately (POST /restart) so the UI can fire it
    immediately after.

    Windows-safe by design — this is the path that was bricking installs:
      * Every git call runs with `-c gc.auto=0` so git never repacks
        objects mid-update (the usual cause of "Unlink of file failed:
        .git/objects/..." when the app process is still running), plus
        GIT_OPTIONAL_LOCKS=0 to stop background index refreshes grabbing
        locks the running process holds.
      * It no longer refuses when tracked files look "modified" — on
        Windows a CRLF checkout makes git report every file as changed,
        which used to abort the update outright.
      * When a clean fast-forward isn't possible (CRLF-dirtied tree, a
        prior half-applied pull, or diverging history) it hard-resets to
        origin/master. That's the same self-healing behaviour the
        installer's Autostart.ps1 / Update.ps1 / recover-windows.ps1 use,
        so a broken checkout repairs itself on the next update or reboot.
    Untracked files (crawl data, logs, the venv) survive the reset."""
    import subprocess as _sp
    repo = os.path.dirname(os.path.abspath(__file__))
    if not os.path.isdir(os.path.join(repo, '.git')):
        return jsonify({'ok': False, 'error': 'Not a git checkout — install via git clone to use auto-update.'}), 400

    _env = {**os.environ, 'GIT_OPTIONAL_LOCKS': '0'}
    def _git(*args, timeout=60):
        return _sp.run(['git', '-C', repo, '-c', 'gc.auto=0', *args],
                       capture_output=True, text=True, timeout=timeout, env=_env)

    def _req_hash():
        """SHA1 of requirements.txt so we can tell if deps changed and need a
        reinstall after the pull — an update that adds a dependency would
        otherwise crash the app on restart and look like a brick."""
        try:
            import hashlib
            with open(os.path.join(repo, 'requirements.txt'), 'rb') as f:
                return hashlib.sha1(f.read()).hexdigest()
        except Exception:
            return ''

    before = _local_commit_sha()
    before_req = _req_hash()
    try:
        fetch = _git('fetch', '--quiet', 'origin', 'master')
        if fetch.returncode != 0:
            return jsonify({'ok': False, 'error': ('git fetch failed: ' + (fetch.stderr or fetch.stdout).strip())[:400]}), 500

        # Prefer a clean fast-forward so untouched installs aren't reset; fall
        # back to a hard reset only when that's impossible (the broken-Windows
        # case). Either way the tree ends up exactly on origin/master.
        ff = _git('merge', '--ff-only', 'origin/master')
        if ff.returncode != 0:
            reset = _git('reset', '--hard', 'origin/master')
            if reset.returncode != 0:
                return jsonify({'ok': False, 'error': ('git reset failed: ' + (reset.stderr or reset.stdout).strip())[:400]}), 500
            msg = 'Hard-reset to origin/master (fast-forward was not possible).'
        else:
            msg = (ff.stdout.strip() or 'Fast-forwarded to origin/master.')

        # Reinstall deps if requirements.txt changed, using this venv's pip,
        # before the restart picks up the new code.
        if _req_hash() != before_req:
            if os.name == 'nt':
                pip = os.path.join(repo, 'venv', 'Scripts', 'pip.exe')
            else:
                pip = os.path.join(repo, 'venv', 'bin', 'pip')
            if os.path.exists(pip):
                dep = _sp.run([pip, 'install', '-r', os.path.join(repo, 'requirements.txt')],
                              capture_output=True, text=True, timeout=300)
                msg += ' Dependencies updated.' if dep.returncode == 0 \
                    else ' WARNING: dependency reinstall failed — restart may fail.'
            else:
                msg += ' (requirements changed but venv pip not found — reinstall manually.)'

        after = _local_commit_sha()
        return jsonify({
            'ok': True,
            'before': before,
            'after': after,
            'changed': before != after,
            'message': msg[:400],
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)[:300]}), 500


@app.route('/restart', methods=['POST'])
def restart_self():
    """Detached restart of the running Flask process. Platform split:
    POSIX uses bash + nohup; Windows uses cmd.exe + start to launch a
    detached pythonw.exe so the new process survives this one's death."""
    import subprocess as _sp
    import sys as _sys
    repo = os.path.dirname(os.path.abspath(__file__))
    own_pid = os.getpid()

    if _sys.platform.startswith('win'):
        # Spawn a detached PowerShell with CREATE_NO_WINDOW + null handles.
        # This is the ONLY launch method that reliably starts from the
        # no-console pythonw process: DETACHED_PROCESS and every cmd.exe/`start`
        # variant silently no-op here (and `start` in a console-less context is
        # what threw the "Windows cannot find '\\'" shell error). The child
        # survives us being killed — Windows doesn't cascade-kill children.
        CREATE_NO_WINDOW = 0x08000000
        pyw = os.path.join(repo, 'venv', 'Scripts', 'pythonw.exe')
        if not os.path.exists(pyw):
            pyw = _sys.executable
        app_py = os.path.join(repo, 'app.py')
        helper = os.path.join(repo, 'restart-helper.ps1')

        if os.path.exists(helper):
            # restart-helper.ps1 kills us, waits for the port to free, then
            # starts + VERIFIES the app, retrying instead of bricking.
            cmd = ['powershell', '-NoProfile', '-ExecutionPolicy', 'Bypass',
                   '-WindowStyle', 'Hidden', '-File', helper,
                   '-OldPid', str(own_pid), '-Port', '5002']
        else:
            # Fallback (pre-helper checkout): inline PowerShell, still no `start`.
            ps = (
                "Start-Sleep -Seconds 1; "
                "try {{ Stop-Process -Id {pid} -Force -ErrorAction SilentlyContinue }} catch {{}}; "
                "Start-Sleep -Seconds 2; "
                "Start-Process -FilePath '{pyw}' -ArgumentList '\"{app}\"' "
                "-WorkingDirectory '{repo}' -WindowStyle Hidden"
            ).format(pid=own_pid, pyw=pyw, app=app_py, repo=repo)
            cmd = ['powershell', '-NoProfile', '-ExecutionPolicy', 'Bypass',
                   '-WindowStyle', 'Hidden', '-Command', ps]
        try:
            _sp.Popen(cmd, creationflags=CREATE_NO_WINDOW,
                      stdin=_sp.DEVNULL, stdout=_sp.DEVNULL, stderr=_sp.DEVNULL)
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)[:300]}), 500
        return jsonify({'ok': True, 'message': 'Restarting in ~3s'})

    # POSIX
    log_path = os.path.expanduser('~/site-crawler/.restart.log')
    cmd = (
        f"sleep 1 && "
        f"kill {own_pid} 2>/dev/null; sleep 1; "
        f"cd {repo} && nohup python3 app.py >/dev/null 2>&1 &"
    )
    try:
        _sp.Popen(['bash', '-c', cmd], start_new_session=True,
                  stdout=open(log_path, 'a'), stderr=_sp.STDOUT)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)[:300]}), 500
    return jsonify({'ok': True, 'message': 'Restarting in ~2s'})


@app.route('/')
def index():
    return render_template('index.html', v=STATIC_VERSION, build_sha=_local_commit_sha())


# =============================================================================
# CMS detection + per-CMS profiles
# =============================================================================
CMS_PROFILES = {
    'shopify': {
        'label': 'Shopify',
        'exclude_patterns': [
            '*?variant=*', '*&variant=*',
            '*/cart', '*/cart/*',
            '*/account*',
            '*/challenge',
            '*/policies/*',
            '*srsltid=*',
            '*/cdn/shop/*',
            '*/collections/*/products/*',  # duplicate products nested under collections
        ],
        'suggested_settings': {'render_js': False, 'max_workers': 5},
        'schema_warnings': [
            'OnlineStore schema without a visible physical address triggers Google validation warnings — prefer plain Organization.',
            'Shopify themes often auto-emit Organization + WebSite schema; check for duplication before adding custom blocks.',
        ],
        'tips': [
            'Shopify uses /search?q= — valid SearchAction target.',
            'Watch for ?variant= and ?srsltid= duplicate URLs.',
        ],
    },
    'wordpress': {
        'label': 'WordPress',
        'exclude_patterns': [
            '*/wp-admin/*', '*/wp-json/*', '*/wp-includes/*',
            '*/wp-content/uploads/*',
            '*/feed/*', '*/feed', '*/?feed=*',
            '*/author/*', '*/tag/*',
            '*/comments/feed*', '*/trackback*',
            '*?p=*', '*?replytocom=*', '*?unapproved=*',
            '*/xmlrpc.php',
        ],
        # 1.0s delay + 3 workers = ~3 req/s, comfortably under Wordfence's default
        # "more than 240 req/min" advanced-block threshold even at burst.
        'suggested_settings': {'render_js': False, 'max_workers': 3, 'crawl_delay': 1.0},
        'schema_warnings': [],
        'tips': [
            'WordPress on shared hosting can throttle under load — reduce workers to 2-3 on fragile sites.',
            'Author, tag, and feed URLs rarely deserve indexing.',
        ],
    },
    'wordpress_yoast': {
        'label': 'WordPress + Yoast SEO',
        'exclude_patterns': [
            '*/wp-admin/*', '*/wp-json/*', '*/wp-includes/*',
            '*/wp-content/uploads/*',
            '*/feed/*', '*/feed', '*/?feed=*',
            '*/author/*', '*/tag/*',
            '*/comments/feed*', '*/trackback*',
            '*?p=*', '*?replytocom=*',
        ],
        'suggested_settings': {'render_js': False, 'max_workers': 3, 'crawl_delay': 1.0},
        'schema_warnings': [
            'Yoast auto-emits an @graph with Organization, WebSite, WebPage, Article, BreadcrumbList. Do NOT duplicate these in custom schema blocks.',
            'Yoast can also emit FAQPage if the FAQ block is used — audit before writing your own FAQPage.',
        ],
        'tips': [
            'If you\'re adding custom schema, put it in a new @graph node with distinct @id values.',
        ],
    },
    'wordpress_rankmath': {
        'label': 'WordPress + Rank Math',
        'exclude_patterns': [
            '*/wp-admin/*', '*/wp-json/*', '*/wp-includes/*',
            '*/wp-content/uploads/*',
            '*/feed/*', '*/feed', '*/?feed=*',
            '*/author/*', '*/tag/*',
            '*/comments/feed*', '*/trackback*',
            '*?p=*', '*?replytocom=*',
        ],
        'suggested_settings': {'render_js': False, 'max_workers': 3, 'crawl_delay': 1.0},
        'schema_warnings': [
            'Rank Math auto-emits Organization, WebSite, WebPage, BreadcrumbList. Check before adding custom blocks.',
        ],
        'tips': [],
    },
    'webflow': {
        'label': 'Webflow',
        'exclude_patterns': [],
        'suggested_settings': {'render_js': False, 'max_workers': 5},
        'schema_warnings': [],
        'tips': [
            'Webflow handles trailing slashes at the server — confirm redirect behaviour is consistent.',
        ],
    },
    'wix': {
        'label': 'Wix',
        'exclude_patterns': [],
        'suggested_settings': {'render_js': True, 'max_workers': 1},
        'schema_warnings': [],
        'tips': [
            'Wix is heavily client-side rendered — enable Render JS or the link graph will be incomplete.',
        ],
    },
    'squarespace': {
        'label': 'Squarespace',
        'exclude_patterns': [],
        'suggested_settings': {'render_js': True, 'max_workers': 3},
        'schema_warnings': [],
        'tips': [
            'Squarespace injects nav client-side — enable Render JS to discover all pages.',
            'Meta title template is "<page title> — <site title>" by default.',
        ],
    },
    'kajabi': {
        'label': 'Kajabi',
        'exclude_patterns': ['*/my-library*', '*/offers/*', '*/checkouts/*'],
        'suggested_settings': {'render_js': False, 'max_workers': 3},
        'schema_warnings': [],
        'tips': [
            'Kajabi lays out testimonials deep in the DOM — review word counts manually for sales pages.',
        ],
    },
    'ghost': {
        'label': 'Ghost',
        'exclude_patterns': ['*/ghost/*', '*/rss/*', '*/amp/*'],
        'suggested_settings': {'render_js': False, 'max_workers': 5},
        'schema_warnings': [],
        'tips': [],
    },
    'drupal': {
        'label': 'Drupal',
        'exclude_patterns': ['*/user/*', '*/node/add/*', '*/taxonomy/*', '*/admin/*'],
        'suggested_settings': {'render_js': False, 'max_workers': 3},
        'schema_warnings': [],
        'tips': [],
    },
    'hubspot': {
        'label': 'HubSpot CMS',
        'exclude_patterns': ['*/_hcms/*', '*/hs-fs/*'],
        'suggested_settings': {'render_js': False, 'max_workers': 5},
        'schema_warnings': [],
        'tips': [],
    },
    'joomla': {
        'label': 'Joomla',
        'exclude_patterns': ['*/administrator/*'],
        'suggested_settings': {'render_js': False, 'max_workers': 3},
        'schema_warnings': [],
        'tips': [],
    },
}


def detect_cms(url, html=None, headers=None):
    """Identify the CMS (and SEO plugin if any) from the HTML + response headers.

    Returns {'cms': key, 'label': str, 'confidence': 'high|medium|low', 'signals': [str]}
    or {'cms': None, ...} if nothing recognisable is found.
    """
    headers = {k.lower(): (v or '') for k, v in (headers or {}).items()}
    html_sample = (html or '')[:80000]  # cap scan to first 80kb — fingerprints live in <head>
    lower = html_sample.lower()
    signals = []
    confidence = 'low'

    def meta_generator():
        m = _re.search(r'<meta[^>]+name=["\']generator["\'][^>]+content=["\']([^"\']+)["\']', html_sample, _re.I)
        return m.group(1) if m else ''

    gen = meta_generator()
    gen_lower = gen.lower()

    # --- High-confidence header/meta matches ---
    if 'x-shopify-stage' in headers or 'x-shopid' in headers or 'cdn.shopify.com' in lower or '/cdn/shop/' in lower:
        signals.append('Shopify CDN / header fingerprint')
        return {'cms': 'shopify', 'label': CMS_PROFILES['shopify']['label'], 'confidence': 'high', 'signals': signals}
    if 'x-wix-request-id' in headers or 'wix.com' in headers.get('x-powered-by', '').lower() or 'wixsite.com' in lower or gen_lower.startswith('wix.com'):
        signals.append('Wix header / generator fingerprint')
        return {'cms': 'wix', 'label': CMS_PROFILES['wix']['label'], 'confidence': 'high', 'signals': signals}
    if 'squarespace' in headers.get('server', '').lower() or 'static1.squarespace.com' in lower or 'squarespace-cdn.com' in lower or gen_lower.startswith('squarespace'):
        signals.append('Squarespace CDN / server fingerprint')
        return {'cms': 'squarespace', 'label': CMS_PROFILES['squarespace']['label'], 'confidence': 'high', 'signals': signals}

    # Webflow
    if 'webflow.com' in lower or _re.search(r'<html[^>]+data-wf-', html_sample) or gen_lower.startswith('webflow'):
        signals.append('Webflow generator / data-wf-* attributes')
        return {'cms': 'webflow', 'label': CMS_PROFILES['webflow']['label'], 'confidence': 'high', 'signals': signals}

    # Kajabi
    if 'kajabi-storefronts-production' in lower or gen_lower.startswith('kajabi'):
        signals.append('Kajabi storefront assets')
        return {'cms': 'kajabi', 'label': CMS_PROFILES['kajabi']['label'], 'confidence': 'high', 'signals': signals}

    # Ghost
    if gen_lower.startswith('ghost') or '/ghost/' in lower or headers.get('x-powered-by', '').lower().startswith('ghost'):
        signals.append('Ghost generator / admin path')
        return {'cms': 'ghost', 'label': CMS_PROFILES['ghost']['label'], 'confidence': 'high', 'signals': signals}

    # HubSpot
    if 'hs-scripts.com' in lower or 'hubspot.net' in lower or 'hubspotusercontent' in lower or gen_lower.startswith('hubspot'):
        signals.append('HubSpot asset CDN')
        return {'cms': 'hubspot', 'label': CMS_PROFILES['hubspot']['label'], 'confidence': 'high', 'signals': signals}

    # Drupal
    if gen_lower.startswith('drupal') or 'x-generator' in headers and 'drupal' in headers['x-generator'].lower() or '/sites/default/files/' in lower:
        signals.append('Drupal generator / asset path')
        return {'cms': 'drupal', 'label': CMS_PROFILES['drupal']['label'], 'confidence': 'high', 'signals': signals}

    # Joomla
    if gen_lower.startswith('joomla'):
        signals.append('Joomla generator meta')
        return {'cms': 'joomla', 'label': CMS_PROFILES['joomla']['label'], 'confidence': 'high', 'signals': signals}

    # --- WordPress detection (+ SEO plugin) ---
    wp_signals = []
    if 'wp-content' in lower or '/wp-includes/' in lower: wp_signals.append('wp-content / wp-includes asset path')
    if '/wp-json/' in lower: wp_signals.append('WordPress REST API link')
    if gen_lower.startswith('wordpress'): wp_signals.append(f'generator: {gen}')
    if headers.get('x-powered-by', '').lower().startswith('wordpress'): wp_signals.append('X-Powered-By: WordPress')
    if wp_signals:
        signals.extend(wp_signals)
        # SEO plugin fingerprints
        if _re.search(r'yoast seo', lower) or 'yoast.com' in lower or _re.search(r'yoast-schema-graph', lower):
            signals.append('Yoast SEO detected')
            return {'cms': 'wordpress_yoast', 'label': CMS_PROFILES['wordpress_yoast']['label'], 'confidence': 'high', 'signals': signals}
        if _re.search(r'rank math', lower) or 'rankmath' in lower:
            signals.append('Rank Math detected')
            return {'cms': 'wordpress_rankmath', 'label': CMS_PROFILES['wordpress_rankmath']['label'], 'confidence': 'high', 'signals': signals}
        return {'cms': 'wordpress', 'label': CMS_PROFILES['wordpress']['label'], 'confidence': 'high', 'signals': signals}

    return {'cms': None, 'label': 'Unknown / custom', 'confidence': 'low', 'signals': []}


@app.route('/detect-cms', methods=['POST'])
def detect_cms_route():
    """Fetch a URL and identify the CMS. Used standalone and also at crawl start."""
    data = request.json or {}
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({'error': 'URL required'}), 400
    if not url.startswith('http'):
        url = 'https://' + url
    try:
        resp = _http_get(url, timeout=10, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/132.0.0.0 Safari/537.36'
        })
        result = detect_cms(url, resp.text, dict(resp.headers))
        if result.get('cms'):
            prof = CMS_PROFILES.get(result['cms'], {})
            result['profile'] = {
                'exclude_patterns': prof.get('exclude_patterns', []),
                'suggested_settings': prof.get('suggested_settings', {}),
                'schema_warnings': prof.get('schema_warnings', []),
                'tips': prof.get('tips', []),
            }
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': f'Detection failed: {str(e)[:100]}'}), 500

_CRAWL_NOISE_PARAMS = frozenset({
    # Tracking
    'fbclid', 'gclid', 'mc_cid', 'mc_eid', 'gad_source', 'gbraid', 'wbraid',
    'msclkid', 'yclid', 'dclid', 'igshid', 'srsltid',
    'ref', 'ref_src', 'ref_url',
    # WooCommerce action endpoints — not real pages.
    'add-to-cart', 'remove_item', 'removed_item', 'undo_item',
    'wc-ajax', 'wc-api', 'wcml_currency', 'orderby', 'product-page',
    'min_price', 'max_price',
    # Other common ecommerce/forum noise
    'replytocom', 'unapproved', 'moderation-hash',
    'share', 'sharesource',
})

# Match a bare email shape in an href that lacks the mailto: prefix.
# Authors sometimes write <a href="sales@example.com"> by mistake — without
# this filter, urljoin resolves it relative to the current page and the
# email lands in the crawl as a fake URL like /contact/sales@example.com.
_MAILTO_NO_SCHEME_RE = _re.compile(r'^[^/\s:?#]+@[^/\s:?#]+\.[A-Za-z]{2,}$')


def _normalize_crawl_url(url):
    """Normalize URL for deduplication: strip fragments, utm/tracking/action
    params, lowercase host.

    Strips WooCommerce action endpoints (?add-to-cart=, ?wc-ajax= …) so
    every product page's "Add to cart" button doesn't surface as its own
    URL with no meta description.
    """
    from urllib.parse import urlparse, urlunparse, parse_qs, urlencode
    parsed = urlparse(url)
    path = parsed.path or '/'
    if parsed.query:
        params = parse_qs(parsed.query, keep_blank_values=True)
        cleaned = {
            k: v for k, v in params.items()
            if not k.lower().startswith('utm_')
            and k.lower() not in _CRAWL_NOISE_PARAMS
        }
        query = urlencode(cleaned, doseq=True)
    else:
        query = ''
    return urlunparse((parsed.scheme.lower(), parsed.netloc.lower(), path, '', query, ''))


def _crawl_slash_alt(url):
    """Return the slash-toggled variant of a URL for dedup, or None if not applicable."""
    from urllib.parse import urlparse, urlunparse
    p = urlparse(url)
    path = p.path
    last_seg = path.rstrip('/').split('/')[-1]
    if '.' in last_seg:
        return None
    if path.endswith('/') and path != '/':
        alt_path = path.rstrip('/')
    else:
        alt_path = path + '/'
    return urlunparse((p.scheme, p.netloc, alt_path, p.params, p.query, p.fragment))


def _detect_js_platform(html):
    """Cheap signature check for JS-rendered platforms. Returns a human-
    readable platform name or None.

    Used to tell the user "this site needs Render JS" instead of firing a
    misleading "No analytics detected" issue when the crawler saw only the
    unhydrated HTML shell.
    """
    if not html:
        return None
    sample = html[:200_000]  # enough for <head> + top of <body>
    checks = [
        ('Wix',         [r'static\.parastorage\.com', r'wix\.com/thunderbolt', r'wixCIDX', r'<html[^>]*wix']),
        ('Shopify',     [r'cdn\.shopify\.com', r'Shopify\.shop', r'shopify-features']),
        ('Squarespace', [r'static1\.squarespace\.com', r'squarespace-cdn\.com', r'Squarespace\.Constants']),
        ('Webflow',     [r'uploads-ssl\.webflow\.com', r'data-wf-page', r'webflow\.js']),
        ('Next.js',     [r'id="__next"', r'__NEXT_DATA__']),
        ('Nuxt',        [r'id="__nuxt"', r'window\.__NUXT__']),
        ('Gatsby',      [r'id="___gatsby"']),
        ('React SPA',   [r'<div[^>]+id="root"[^>]*></div>\s*<script']),
        ('Vue SPA',     [r'<div[^>]+id="app"[^>]*></div>\s*<script']),
    ]
    import re as _re_local
    for name, patterns in checks:
        for pat in patterns:
            if _re_local.search(pat, sample, _re_local.I):
                return name
    return None


# Third-party widget images the site owner cannot meaningfully add alt text
# to — reCAPTCHA badges, analytics 1×1 pixels, chat-widget assets, ad-tech
# beacons. Flagging them as "missing alt" pollutes every page that loads a
# contact form. Match by hostname OR filename — both routes needed because
# self-hosted recaptcha-black.svg in a WP theme bypasses the host check.
# Mirrored from seo-tool/app.py — keep both lists in sync.
_THIRD_PARTY_IMG_HOSTS = (
    'gstatic.com',
    'googletagmanager.com',
    'google-analytics.com',
    'googleadservices.com',
    'doubleclick.net',
    'static.hotjar.com',
    'script.hotjar.com',
    'cdn.intercomcdn.com',
    'js.intercomcdn.com',
    'cdn.intercom.io',
    'widget.crisp.chat',
    'client.crisp.chat',
    'static.cloudflareinsights.com',
    'analytics.tiktok.com',
    'connect.facebook.net',
    'i.pinimg.com',
    'ct.pinterest.com',
    'hs-analytics.net',
    'js.hs-scripts.com',
)
_THIRD_PARTY_IMG_FILE_RE = _re.compile(
    r'(?:/recaptcha[/_\-]|recaptcha[_\-](?:black|white|logo)|/g\.gif$|/pixel\.gif$|'
    r'/spacer\.gif$|/tracking[_\-]pixel|fbq[_\-]pixel|/fb-pixel|/ga-pixel)',
    _re.I,
)


def _is_third_party_widget_image(abs_src):
    """Return True for images injected by third-party widgets/trackers.

    The crawler can see these but the site owner cannot meaningfully add alt
    text — reCAPTCHA logos, analytics 1×1 beacons, chat-widget assets, etc.
    Filtering them out keeps the 'imgs missing alt' view actionable.
    """
    if not abs_src:
        return False
    try:
        from urllib.parse import urlparse as _up
        host = (_up(abs_src).hostname or '').lower()
    except Exception:
        host = ''
    if host:
        for h in _THIRD_PARTY_IMG_HOSTS:
            if host == h or host.endswith('.' + h):
                return True
    if _THIRD_PARTY_IMG_FILE_RE.search(abs_src):
        return True
    return False


# alt="" is the spec-correct pattern for decorative images, but on real sites
# the bucket also catches photos / screenshots / logos where the alt was simply
# forgotten. Match the filename against page-builder shape exports (Elementor,
# Figma, Sketch, XD) and common UI ornament names — anything that does NOT
# match is treated as likely content and surfaced for owner review.
_DECORATIVE_FILENAME_RE = _re.compile(
    r'(?:'
    r'^(?:layer|group|mask[-_]?group|path|vector|rectangle|ellipse|frame|union|subtract|clip|component|line|polygon|oval|artboard)[-_ ]?\d*'
    r'|quotation|quote[-_]?mark'
    r'|(?:^|[-_/])(?:bg|background|backdrop|hero[-_]?bg|pattern|texture|noise|gradient|overlay|stripe|grid|mesh)(?:[-_]|$)'
    r'|(?:^|[-_/])(?:icon|ico|sprite|emoji|emote|bullet|chevron|caret|burger|hamburger|loader|spinner|placeholder|divider|separator|ornament|accent|swirl|squiggle|ribbon)(?:[-_]|$)'
    r'|(?:^|[-_/])(?:star|sparkle|shape|blob|leaf|petal|circle|square|triangle)(?:[-_]|$)'
    r')',
    _re.I,
)


def _filename_looks_decorative(src):
    """Heuristic: does this image filename look like a decorative shape /
    icon / pattern export rather than a content photo / screenshot / logo?

    Decorative (True):  Layer-1.svg, Group-83.png, Mask-Group-2.png,
                        quotation.png, pattern-dots.svg, hero-bg.jpg,
                        icon-arrow.svg, divider.png
    Content (False):    hero-team.jpg, screenshot-2026-04-12.png,
                        logo-white.png, unrecognizable-man-working.jpg

    Used to split <img alt=""> into 'genuinely decorative' (info) and
    'likely content with forgotten alt' (surfaced as an issue).
    """
    if not src:
        return True
    try:
        from urllib.parse import urlparse as _up
        path = _up(src).path or src
    except Exception:
        path = src
    fname = path.rsplit('/', 1)[-1]
    if not fname:
        return True
    fname = _re.sub(r'\.(svg|png|jpe?g|gif|webp|avif|ico|bmp)$', '', fname, flags=_re.I)
    fname = _re.sub(r'[-_]\d{2,4}x\d{2,4}$', '', fname)
    fname = _re.sub(r'@\d+x$', '', fname)
    fname = _re.sub(r'[-_]scaled$', '', fname, flags=_re.I)
    return bool(_DECORATIVE_FILENAME_RE.search(fname))


def _parse_no_js_subset(html, base_url):
    """Parse a subset of SEO-critical fields from raw HTML — used for the
    JS-vs-non-JS comparison so we can show what content is missing when JS
    isn't executed. Same parsers/idioms as the main pass in `_crawl_page`,
    just trimmed to the fields we diff against. Kept self-contained so the
    seo-tool port stays a copy-paste.
    """
    from urllib.parse import urlparse, urljoin
    out = {
        'title': '', 'meta_description': '', 'h1': '', 'h1_count': 0,
        'word_count': 0, 'schema_types': [],
        'internal_links_count': 0, 'external_links_count': 0,
        'images_count': 0, 'images_no_alt': 0,
    }
    try:
        soup = BeautifulSoup(html or '', 'html.parser')
        t = soup.find('title')
        out['title'] = t.get_text(strip=True) if t else ''
        m = soup.find('meta', attrs={'name': 'description'})
        out['meta_description'] = m.get('content', '') if m else ''
        import re as _re_h
        def _clean_h(tag):
            txt = ' '.join(tag.get_text(' ', strip=True).split())
            return _re_h.sub(r'\s+([,.;:!?\)\]])', r'\1', txt)
        h1s = [_clean_h(h) for h in soup.find_all('h1')]
        h1s = [h for h in h1s if h]
        out['h1'] = h1s[0] if h1s else ''
        out['h1_count'] = len(h1s)
        for s in soup(['script', 'style', 'noscript']):
            s.decompose()
        body_text = ' '.join(soup.get_text(separator=' ', strip=True).split())
        out['word_count'] = len(body_text.split()) if body_text else 0
        types = []
        def _push(v):
            if isinstance(v, list):
                for x in v:
                    if isinstance(x, str): types.append(x)
            elif isinstance(v, str):
                types.append(v)
        for sc in soup.find_all('script', attrs={'type': 'application/ld+json'}):
            try:
                ld = json.loads(sc.string or '')
                if isinstance(ld, dict):
                    if '@type' in ld: _push(ld['@type'])
                    if '@graph' in ld:
                        for it in ld['@graph']:
                            if isinstance(it, dict) and '@type' in it:
                                _push(it['@type'])
                elif isinstance(ld, list):
                    for it in ld:
                        if isinstance(it, dict) and '@type' in it:
                            _push(it['@type'])
            except Exception:
                pass
        out['schema_types'] = types
        try:
            host = (urlparse(base_url).netloc or '').lower().replace('www.', '')
        except Exception:
            host = ''
        for a in soup.find_all('a', href=True):
            href = a.get('href', '').strip()
            if not href or href.startswith(('#', 'javascript:', 'mailto:', 'tel:')):
                continue
            try:
                abs_url = urljoin(base_url, href)
                link_host = (urlparse(abs_url).netloc or '').lower().replace('www.', '')
            except Exception:
                continue
            if link_host == host or not link_host:
                out['internal_links_count'] += 1
            else:
                out['external_links_count'] += 1
        imgs = soup.find_all('img')
        out['images_count'] = len(imgs)
        out['images_no_alt'] = sum(1 for i in imgs if not (i.get('alt') or '').strip())
    except Exception:
        pass
    return out


def _compute_js_diff(js, nojs):
    """Compare rendered (post-JS) fields against pre-JS subset and classify
    severity. See seo-tool/app.py:_compute_js_diff for the full rationale.
    Severity ladder: critical (title/meta/schema), high (h1/word_count),
    medium (links/images), none (page is server-rendered correctly).
    """
    fields_differ = []
    sev = 'none'
    rendered_title = (js.get('title') or '').strip()
    nojs_title = (nojs.get('title') or '').strip()
    if rendered_title != nojs_title:
        fields_differ.append('title'); sev = 'critical'
    rendered_meta = (js.get('meta_description') or '').strip()
    nojs_meta = (nojs.get('meta_description') or '').strip()
    if rendered_meta != nojs_meta:
        fields_differ.append('meta_description'); sev = 'critical'
    js_schema = sorted(set(js.get('schema_types') or []))
    nojs_schema = sorted(set(nojs.get('schema_types') or []))
    if js_schema != nojs_schema:
        fields_differ.append('schema_types'); sev = 'critical'
    rendered_h1 = (js.get('h1') or '').strip()
    nojs_h1 = (nojs.get('h1') or '').strip()
    if rendered_h1 != nojs_h1:
        fields_differ.append('h1')
        if sev != 'critical': sev = 'high'
    js_wc = js.get('word_count') or 0
    nojs_wc = nojs.get('word_count') or 0
    if js_wc > 0 and abs(js_wc - nojs_wc) / max(js_wc, 1) > 0.25:
        fields_differ.append('word_count')
        if sev != 'critical': sev = 'high'
    js_il = js.get('internal_links') or 0
    nojs_il = nojs.get('internal_links_count') or 0
    if js_il > 0 and abs(js_il - nojs_il) / max(js_il, 1) > 0.25:
        fields_differ.append('internal_links')
        if sev not in ('critical', 'high'): sev = 'medium'
    js_el = js.get('external_links') or 0
    nojs_el = nojs.get('external_links_count') or 0
    if js_el > 0 and abs(js_el - nojs_el) / max(js_el, 1) > 0.25:
        fields_differ.append('external_links')
        if sev not in ('critical', 'high'): sev = 'medium'
    js_imgs = js.get('images_total') or 0
    nojs_imgs = nojs.get('images_count') or 0
    if js_imgs > 0 and abs(js_imgs - nojs_imgs) / max(js_imgs, 1) > 0.10:
        fields_differ.append('images_total')
        if sev not in ('critical', 'high'): sev = 'medium'
    if (js.get('images_no_alt') or 0) != (nojs.get('images_no_alt') or 0):
        fields_differ.append('images_no_alt')
        if sev not in ('critical', 'high'): sev = 'medium'
    return {'severity': sev, 'fields': fields_differ}


def _detect_waf_block(resp):
    """Sniff a 403/429/503 response for WAF block-page markers.
    Returns 'wordfence' / 'cloudflare' / 'sucuri' / 'siteground' / None.
    Retrying these against the same IP just escalates the block — the caller
    should bail out of retries and let the host-pause kick in for ~5 minutes."""
    if resp is None: return None
    if resp.status_code not in (403, 429, 503): return None
    try:
        body_head = (resp.text or '')[:4000].lower()
    except Exception:
        return None
    if 'wordfence' in body_head or 'generated by wordfence' in body_head:
        return 'wordfence'
    if 'attention required' in body_head and 'cloudflare' in body_head:
        return 'cloudflare'
    if 'sucuri' in body_head and ('website firewall' in body_head or 'access denied' in body_head):
        return 'sucuri'
    if 'siteground' in body_head and ('blocked' in body_head or 'security' in body_head):
        return 'siteground'
    return None


def _crawl_page(url, session, domain, pw_page=None, ignore_noindex=False, capture_no_js=False):
    """Crawl a single page and return audit data dict.

    If ``pw_page`` (a live Playwright page) is provided, the HTML body will be
    re-fetched via a headless browser so JS-rendered content is captured. We
    still do the initial ``requests.get`` to get response headers / redirect
    history cheaply and reliably.

    ``capture_no_js``: when True AND ``pw_page`` produced a successful render,
    also parse the original (pre-JS) HTML and attach a subset of fields to
    ``result['non_js']`` plus a diff/severity summary at ``result['js_diff']``.
    Lets the user see what content is JS-only and therefore at risk for AI
    crawlers (GPTBot, ClaudeBot etc.) which mostly don't execute JS.
    """
    from urllib.parse import urlparse, urljoin
    result = {
        'url': url, 'status_code': 0, 'content_type': '', 'last_modified': '', 'response_time': 0,
        'title': '', 'title_len': 0, 'meta_description': '', 'meta_len': 0,
        'h1': '', 'h1_list': [], 'h2_list': [], 'h2_count': 0,
        'canonical': '', 'canonical_match': False, 'canonical_kind': None,
        'word_count': 0, 'internal_links': 0, 'external_links': 0,
        'internal_link_urls': [], 'images_total': 0, 'images_no_alt': 0,
        'schema_types': [], 'indexable': True, 'is_pagination': False, 'issues': [], 'error': None,
        'depth': 0, 'redirect_url': None, 'redirect_kind': None,
        'redirect_status': None, 'redirect_hops': 0, 'redirect_chain': [],
        'body_hash': '', 'security': {}, 'mixed_content': [],
        'url_issues': [], 'hreflang': [], 'x_robots_tag': '',
        'og_tags': {}, 'twitter_tags': {}, 'analytics': [],
    }

    try:
        # Retry loop for transient failures (5xx, 429, connection errors).
        # 403 gets a UA-fallback instead of a delay-and-retry, because 403 is
        # usually bot-fingerprint detection rather than a rate limit.
        resp = None
        retries_done = 0
        last_exc = None
        ssl_bypassed = False
        for attempt in range(3):  # 1 primary + 2 retries
            try:
                resp = session.get(url, timeout=15, allow_redirects=True)
                last_exc = None
            except requests.exceptions.SSLError as e:
                # Broken/incomplete certificate chain (most often a missing
                # intermediate cert). The site IS reachable — browsers fetch the
                # missing intermediate via AIA automatically, requests does not.
                # Retry with verification OFF so the crawl can proceed, flag it
                # loudly, and disable verify for the rest of this session so we
                # don't pay a failed handshake on every subsequent page.
                last_exc = e
                try:
                    import urllib3
                    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
                except Exception:
                    pass
                try:
                    resp = session.get(url, timeout=15, allow_redirects=True, verify=False)
                    last_exc = None
                    ssl_bypassed = True
                    session.verify = False
                except requests.exceptions.RequestException as e2:
                    last_exc = e2
                    resp = None
            except requests.exceptions.RequestException as e:
                last_exc = e
                resp = None
            # Decide whether to retry, fall back UA, or accept the response.
            if resp is not None:
                if resp.status_code == 429:
                    # Respect Retry-After header if present, else exponential wait.
                    ra = resp.headers.get('Retry-After', '').strip()
                    try: wait = float(ra) if ra else 2 * (2 ** attempt)
                    except ValueError: wait = 2 * (2 ** attempt)
                    wait = min(wait, 10)
                    if attempt < 2:
                        retries_done += 1
                        time.sleep(wait)
                        continue
                elif 500 <= resp.status_code < 600:
                    # WAF block-page check — Wordfence/Cloudflare/Sucuri 503s.
                    # Retrying just escalates the block, so bail and let the
                    # host-pause window cool the IP down.
                    _waf = _detect_waf_block(resp)
                    if _waf:
                        result['_waf_block'] = _waf
                        break
                    if attempt < 2:
                        retries_done += 1
                        time.sleep(0.5 * (2 ** attempt))  # 0.5s, 1s, 2s
                        continue
                elif resp.status_code == 403:
                    # WAF check first — UA-swap won't help against an IP-flag,
                    # and trying it just spends quota.
                    _waf = _detect_waf_block(resp)
                    if _waf:
                        result['_waf_block'] = _waf
                        break
                    # One UA-swap attempt to dodge Cloudflare-style fingerprinting.
                    # Only swap UA on the first attempt so we don't keep cycling.
                    if attempt == 0:
                        for _alt_ua in (
                            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36',
                            'Mozilla/5.0 (X11; Linux x86_64; rv:128.0) Gecko/20100101 Firefox/128.0',
                        ):
                            try:
                                alt = session.get(url, timeout=15, allow_redirects=True,
                                                  headers={'User-Agent': _alt_ua})
                                if alt.status_code not in (403, 429):
                                    resp = alt
                                    result['ua_fallback'] = _alt_ua.split(')')[0] + ')'
                                    break
                            except Exception:
                                continue
                # Response is either final (2xx/3xx/4xx non-403/429) or exhausted retries.
                break
            # resp is None → transient connection error → retry with backoff
            if attempt < 2:
                retries_done += 1
                time.sleep(0.5 * (2 ** attempt))
                continue
        if resp is None:
            # All retries exhausted
            raise last_exc if last_exc else requests.exceptions.RequestException('Connection failed after retries')
        result['retries'] = retries_done
        if ssl_bypassed:
            result['ssl_verify_failed'] = True
            result['issues'].append('SSL certificate verification failed (incomplete chain / untrusted cert) — crawled with verification disabled; fix the cert chain')
        result['status_code'] = resp.status_code
        result['response_time'] = round(resp.elapsed.total_seconds(), 2)
        result['content_type'] = resp.headers.get('Content-Type', '')[:50]
        result['last_modified'] = resp.headers.get('Last-Modified', '')[:60]

        # Track redirects — classify by type so trivial normalizations (trailing slash,
        # www, https) don't pollute the main Redirect bucket
        if resp.history:
            from urllib.parse import urlparse as _up
            orig = _up(url)
            final = _up(resp.url)
            hops = len(resp.history)
            hop_lbl = f'{hops} hop{"s" if hops > 1 else ""}'
            result['redirect_hops'] = hops
            result['redirect_chain'] = [
                {'url': h.url, 'status': h.status_code} for h in resp.history
            ] + [{'url': resp.url, 'status': resp.status_code}]

            # No-op loop: the server bounced us through 1+ hops but landed
            # right back at the requested URL (same scheme + host + path +
            # query). Common with Wordfence / SiteGround / Cloudflare cookie
            # handshakes. Not a real redirect from the user/SEO perspective —
            # don't set redirect_url, don't append an issue, don't classify.
            # Otherwise the Redirects bucket showed "From: X  To: X" rows
            # the user couldn't action.
            is_noop_loop = (
                orig.scheme == final.scheme
                and orig.netloc.lower() == final.netloc.lower()
                and orig.path == final.path
                and orig.query == final.query
            )

            if not is_noop_loop:
                result['redirect_url'] = resp.url
                # The first hop carries the real 3xx code (301/302/307/308).
                # status_code on the row is the FINAL 200 because we follow
                # redirects, so capture the originating status separately —
                # otherwise the Redirects view can't tell permanent from
                # temporary.
                result['redirect_status'] = resp.history[0].status_code

                same_host = orig.netloc.lstrip('www.') == final.netloc.lstrip('www.')
                same_path_stripped = orig.path.rstrip('/') == final.path.rstrip('/')
                same_qs = orig.query == final.query

                redirect_kind = 'other'
                if same_host and same_path_stripped and same_qs:
                    if orig.scheme != final.scheme and orig.scheme == 'http':
                        redirect_kind = 'http_to_https'
                        result['issues'].append(f'HTTP→HTTPS redirect ({hop_lbl})')
                    elif orig.netloc != final.netloc:
                        redirect_kind = 'www_normalize'
                        result['issues'].append(f'www normalization redirect ({hop_lbl})')
                    elif orig.path != final.path:
                        # Pure trailing slash difference
                        if final.path == orig.path + '/' or orig.path == final.path + '/':
                            redirect_kind = 'trailing_slash'
                            result['issues'].append(f'Trailing slash redirect ({hop_lbl})')
                        else:
                            redirect_kind = 'other'
                            result['issues'].append(f'Redirect ({hop_lbl})')
                    else:
                        redirect_kind = 'other'
                        result['issues'].append(f'Redirect ({hop_lbl})')
                else:
                    result['issues'].append(f'Redirect ({hop_lbl})')

                result['redirect_kind'] = redirect_kind

                # Canonicalize: the row should represent the URL that actually
                # returned 200, not the URL we requested. Without this, the seed
                # AND the redirect target both end up as separate "page" rows when
                # the target is later discovered via an internal link.
                result['url'] = resp.url
                result['original_url'] = url

        # --- URL issues (from the source URL, not redirect target) ---
        from urllib.parse import urlparse as _urlparse2
        _parsed_url = _urlparse2(url)
        _url_path_query = _parsed_url.path + (('?' + _parsed_url.query) if _parsed_url.query else '')
        _url_issues = []
        if any(c.isupper() for c in _parsed_url.path):
            _url_issues.append('uppercase')
        if '_' in _parsed_url.path:
            _url_issues.append('underscores')
        if ' ' in url or '%20' in _parsed_url.path:
            _url_issues.append('contains space')
        if len(url) > 115:
            _url_issues.append(f'over 115 chars ({len(url)})')
        if '//' in _parsed_url.path:
            _url_issues.append('multiple slashes')
        try:
            url.encode('ascii')
        except UnicodeEncodeError:
            _url_issues.append('non-ASCII characters')
        if _parsed_url.query:
            _url_issues.append('contains parameters')
            if any(p in _parsed_url.query for p in ('utm_', 'gclid=', 'fbclid=')):
                _url_issues.append('tracking parameters')
        result['url_issues'] = _url_issues

        # --- Pagination detection ---
        import re as _re_pag
        _pag_path = _re_pag.search(r'/page/\d+/?$', _parsed_url.path)
        _pag_query = _re_pag.search(r'\b(page|paged|pg)\s*=\s*[2-9]\d*', _parsed_url.query)
        result['is_pagination'] = bool(_pag_path or _pag_query)

        # --- Security headers (applies to every response, HTML or not) ---
        hdrs = {k.lower(): v for k, v in resp.headers.items()}
        sec = {
            'hsts': 'strict-transport-security' in hdrs,
            'csp': 'content-security-policy' in hdrs,
            'x_content_type_options': 'x-content-type-options' in hdrs,
            'x_frame_options': 'x-frame-options' in hdrs,
            'referrer_policy': 'referrer-policy' in hdrs,
            'is_https': _parsed_url.scheme == 'https',
        }
        result['security'] = sec
        result['x_robots_tag'] = hdrs.get('x-robots-tag', '')
        if result['x_robots_tag'] and 'noindex' in result['x_robots_tag'].lower():
            result['indexable'] = False

        if resp.status_code >= 400:
            result['error'] = f'HTTP {resp.status_code}'
            result['issues'].append(f'HTTP {resp.status_code} error')
            return result

        # Skip non-HTML
        ctype = result['content_type'].lower()
        if 'text/html' not in ctype and 'application/xhtml' not in ctype:
            result['issues'].append(f'Non-HTML ({ctype.split(";")[0]})')
            return result

        # Limit body size
        raw_html = resp.text[:5_000_000]
        # Keep pre-JS HTML for the JS-vs-non-JS diff (cheap; one extra ref).
        pre_js_html = raw_html
        result['js_rendered'] = False

        # Optional: re-render with Playwright to capture JS-inserted content.
        # Strategy: wait for `load` (fires after all initial subresources), then
        # try (but don't require) networkidle as a short grace window so late
        # analytics scripts (GA4, GTM, FB Pixel) can inject. Wix/Shopify/React
        # SPAs often ping telemetry continuously, so networkidle never settles —
        # we still read .content() regardless so we don't fall back to the
        # empty pre-JS HTML. That silent fallback was why "No analytics
        # detected" fired on JS-rendered sites like Wix.
        if pw_page is not None and ('text/html' in ctype or 'application/xhtml' in ctype):
            try:
                pw_page.goto(resp.url, wait_until='load', timeout=20000)
            except Exception as e:
                # Even if goto times out, the page may have loaded enough of
                # the DOM to be useful — keep going and let content() decide.
                result.setdefault('render_errors', []).append(f'goto: {str(e)[:160]}')
            # Best-effort settle window for late-injected trackers
            try:
                pw_page.wait_for_load_state('networkidle', timeout=4000)
            except Exception:
                pass
            try:
                rendered = pw_page.content()
                if rendered and len(rendered) > 100:
                    raw_html = rendered[:5_000_000]
                    result['js_rendered'] = True
            except Exception as e:
                result.setdefault('render_errors', []).append(f'content: {str(e)[:160]}')

        soup = BeautifulSoup(raw_html, 'html.parser')

        # Title
        title_tag = soup.find('title')
        result['title'] = title_tag.get_text(strip=True) if title_tag else ''
        result['title_len'] = len(result['title'])

        # Meta description
        meta_tag = soup.find('meta', attrs={'name': 'description'})
        result['meta_description'] = meta_tag.get('content', '') if meta_tag else ''
        result['meta_len'] = len(result['meta_description'])

        # H1s and H2s (full lists, not just first/count). Pick the first
        # non-empty H1 for the displayed column value — themes that wrap a
        # logo image in <h1> emit an empty H1 as the first tag, which would
        # otherwise mask a real H1 later in the DOM and cause the row to be
        # flagged as both Missing H1 and Multiple H1s.
        # Pass separator=' ' so inline children (spans, <br>, etc.) don't collide
        # into "connectspeople,pay" when the visible H1 reads "connects people, pay".
        # Webflow/Framer split headings across spans for animated reveals.
        # Also strip the space the separator adds before punctuation tokens.
        import re as _re
        def _clean_heading(tag):
            txt = ' '.join(tag.get_text(' ', strip=True).split())
            return _re.sub(r'\s+([,.;:!?\)\]])', r'\1', txt)
        h1_tags = soup.find_all('h1')
        result['h1_list'] = [_clean_heading(t)[:200] for t in h1_tags]
        result['h1'] = next((t for t in result['h1_list'] if t), '')
        h2_tags = soup.find_all('h2')
        result['h2_list'] = [_clean_heading(t)[:200] for t in h2_tags][:20]
        result['h2_count'] = len(h2_tags)

        # Canonical — classify as self / canonicalised / mismatch
        can_tag = soup.find('link', attrs={'rel': 'canonical'})
        result['canonical'] = can_tag.get('href', '') if can_tag else ''
        if result['canonical']:
            can_abs = urljoin(url, result['canonical'])
            same = can_abs.rstrip('/') == url.rstrip('/') or can_abs.rstrip('/') == resp.url.rstrip('/')
            result['canonical_match'] = same
            if same:
                result['canonical_kind'] = 'self'
            else:
                # canonical points elsewhere — that's a canonicalised page
                result['canonical_kind'] = 'canonicalised'
        else:
            result['canonical_match'] = False
            result['canonical_kind'] = 'missing'

        # Hreflang annotations
        for ln in soup.find_all('link', attrs={'rel': 'alternate'}):
            hl = ln.get('hreflang')
            hf = ln.get('href')
            if hl and hf:
                result['hreflang'].append({'lang': hl, 'href': urljoin(url, hf)})

        # Schema types — schema.org @type can be a single string or a list
        # ("@type": ["Service", "LocalBusiness"]). Flatten to individual
        # strings so the client receives list[str] and rendering doesn't
        # choke on a nested array element.
        def _push_type(val):
            if isinstance(val, list):
                for v in val:
                    if isinstance(v, str):
                        result['schema_types'].append(v)
            elif isinstance(val, str):
                result['schema_types'].append(val)
        for script in soup.find_all('script', attrs={'type': 'application/ld+json'}):
            try:
                ld = json.loads(script.string or '')
                if isinstance(ld, dict):
                    if '@type' in ld:
                        _push_type(ld['@type'])
                    if '@graph' in ld:
                        for item in ld['@graph']:
                            if isinstance(item, dict) and '@type' in item:
                                _push_type(item['@type'])
                elif isinstance(ld, list):
                    for item in ld:
                        if isinstance(item, dict) and '@type' in item:
                            _push_type(item['@type'])
            except Exception:
                pass

        # Open Graph and Twitter Card tags
        for m in soup.find_all('meta'):
            prop = m.get('property', '') or ''
            nm = m.get('name', '') or ''
            content = m.get('content', '') or ''
            if prop.startswith('og:') and content:
                result['og_tags'][prop[3:]] = content[:500]
            elif nm.startswith('twitter:') and content:
                result['twitter_tags'][nm[8:]] = content[:500]

        # Analytics / tracking pixels — inspect script srcs + inline JS
        # Each entry: (label, list of regex patterns to search anywhere in the HTML)
        _TRACKERS = [
            ('GA4',               [r'gtag/js\?id=G-', r"gtag\(\s*'config'\s*,\s*'G-"]),
            ('Universal Analytics', [r'google-analytics\.com/analytics\.js', r"'UA-\d+", r'ua-\d+-\d+']),
            ('Google Tag Manager', [r'googletagmanager\.com/gtm\.js', r"'GTM-[A-Z0-9]+'"]),
            ('Google Ads',        [r'googleadservices\.com/pagead/conversion', r'AW-\d+']),
            ('Facebook Pixel',    [r'connect\.facebook\.net/[^"\']*/fbevents\.js', r"fbq\s*\(\s*['\"]init"]),
            ('TikTok Pixel',      [r'analytics\.tiktok\.com/i18n/pixel']),
            ('LinkedIn Insight',  [r'snap\.licdn\.com/li\.lms-analytics']),
            ('Hotjar',            [r'static\.hotjar\.com/c/hotjar', r'hjSiteSettings', r'\(h,o,t,j,a,r\)']),
            ('Microsoft Clarity', [r'clarity\.ms/tag']),
            ('Mixpanel',          [r'cdn\.mxpnl\.com', r'mixpanel\.init']),
            ('Segment',           [r'cdn\.segment\.com/analytics', r'analytics\.load']),
            ('HubSpot',           [r'js\.hs-scripts\.com', r'js\.hs-analytics\.net']),
            ('Plausible',         [r'plausible\.io/js']),
            ('Fathom',            [r'cdn\.usefathom\.com']),
            ('Matomo/Piwik',      [r'matomo\.php', r'_paq\.push']),
            ('Crazy Egg',         [r'script\.crazyegg\.com']),
            ('Microsoft Ads UET', [r'bat\.bing\.com/bat\.js']),
            ('Pinterest Tag',     [r'pintrk\s*\(\s*[\'"]load']),
            ('Snapchat Pixel',    [r'sc-static\.net/scevent']),
        ]
        # Only scan the raw HTML once; much cheaper than re-traversing soup per pattern
        for label, patterns in _TRACKERS:
            for pat in patterns:
                if _re.search(pat, raw_html, _re.I):
                    result['analytics'].append(label)
                    break

        # Indexability
        robots_tag = soup.find('meta', attrs={'name': _re.compile(r'^robots$', _re.I)})
        robots_content = robots_tag.get('content', '').lower() if robots_tag else ''
        if 'noindex' in robots_content:
            result['indexable'] = False
            result['issues'].append('noindex')

        # Word count — strip nav/footer/script/style + class-based nav for
        # non-semantic sites (Elementor/Divi/WP themes that render nav inside
        # <div class="elementor-nav-menu">). Also prefer <main>/<article>
        # content when present so the body word count reflects actual content,
        # not menu/footer boilerplate (otherwise thin-content detection misfires).
        soup_body = BeautifulSoup(raw_html, 'html.parser')
        for tag in soup_body(['script', 'style', 'nav', 'footer', 'header', 'iframe', 'noscript', 'form', 'svg']):
            tag.decompose()
        try:
            nav_like_re = _re.compile(
                r'(main[-_]?menu|primary[-_]?menu|site[-_]?nav|header[-_]?nav|'
                r'mega[-_]?menu|nav[-_]?menu|top[-_]?menu|mobile[-_]?menu|sub[-_]?menu|'
                r'breadcrumb|footer[-_]?menu|site[-_]?header|site[-_]?footer|'
                r'menu[-_]?wrap|navbar|navigation|'
                r'elementor-nav-menu|elementor-menu|elementor-widget-nav-menu|'
                r'et_pb_menu|divi-menu|menu-item-has-children|\bmenu\b|'
                r'offcanvas|hamburger|dropdown-menu)',
                _re.I,
            )
            role_nav_re = _re.compile(r'^(navigation|menubar|menu)$', _re.I)
            for el in list(soup_body.find_all(['div', 'section', 'ul', 'aside'])):
                if not el.parent:
                    continue
                classes = ' '.join(el.get('class') or [])
                ident = el.get('id') or ''
                role = el.get('role') or ''
                aria = el.get('aria-label') or ''
                if (nav_like_re.search(classes) or nav_like_re.search(ident)
                        or role_nav_re.match(role) or nav_like_re.search(aria)):
                    el.decompose()
        except Exception:
            pass
        # Pick the first <article> that ISN'T a post-grid / related-posts card.
        # Elementor and many blog themes render recent/related posts as
        # <article class="elementor-post elementor-grid-item ...">; grabbing the
        # first of those captures the same boilerplate snippet on every page, so
        # every post looks like an exact body duplicate. Skip those cards.
        def _real_article(sb):
            for art in sb.find_all('article'):
                cls = ' '.join(art.get('class') or [])
                if _re.search(r'elementor-post|elementor-grid-item|elementor-posts|post-grid|related|recent[-_]?post|widget', cls, _re.I):
                    continue
                return art
            return None

        main_container = (
            soup_body.find('main')
            or soup_body.find(attrs={'role': 'main'})
            # Elementor Theme Builder renders the real post body in this widget
            # and ships no <main>/<article> wrapper, so it must come before the
            # generic <article> fallback (which would otherwise hit a grid card).
            or soup_body.find(attrs={'class': _re.compile(r'elementor-widget-theme-post-content', _re.I)})
            or _real_article(soup_body)
            or soup_body.find(id=_re.compile(r'^(main|content|primary|page-content)$', _re.I))
            or soup_body.find(attrs={'class': _re.compile(r'(^|\s)(main-content|page-content|entry-content|post-content|article-content|site-main)(\s|$)', _re.I)})
        )
        text_root = main_container if main_container else soup_body
        body_text = text_root.get_text(separator=' ', strip=True)
        result['word_count'] = len(body_text.split()) if body_text else 0
        # Cap body text at 30k chars — sufficient for shingle-based near-dup
        # detection on any reasonable page; keeps the SSE payload bounded.
        result['body_text'] = (body_text[:30_000] if body_text else '')

        # Body hash for exact-duplicate detection — normalize whitespace first
        import hashlib as _hashlib
        _norm = ' '.join(body_text.lower().split())
        result['body_hash'] = _hashlib.md5(_norm.encode('utf-8', errors='ignore')).hexdigest() if _norm else ''

        # Mixed content: HTTPS page loading HTTP resources.
        # <link> is rel-dependent — most rels are pure metadata (rel="profile"
        # for XFN, rel="canonical", rel="alternate", rel="EditURI"/"pingback"/
        # "https://api.w.org/" for WP) and don't trigger any fetch. Browsers
        # only fire mixed-content warnings on the rels below.
        _MIXED_LINK_RELS = {
            'stylesheet',
            'preload', 'prefetch', 'modulepreload',
            'icon', 'shortcut icon', 'apple-touch-icon',
            'apple-touch-icon-precomposed', 'mask-icon', 'fluid-icon',
            'manifest',
        }
        if result.get('security', {}).get('is_https'):
            mixed = []
            for tag_name, attr in (('img', 'src'), ('script', 'src'),
                                    ('iframe', 'src'), ('video', 'src'),
                                    ('audio', 'src'), ('source', 'src')):
                for t in soup.find_all(tag_name, attrs={attr: True}):
                    v = (t.get(attr) or '').strip()
                    if v.startswith('http://'):
                        mixed.append(v)
            for t in soup.find_all('link', attrs={'href': True}):
                v = (t.get('href') or '').strip()
                if not v.startswith('http://'):
                    continue
                rels = [r.lower() for r in (t.get('rel') or [])]
                if any(r in _MIXED_LINK_RELS for r in rels):
                    mixed.append(v)
            result['mixed_content'] = mixed[:20]

        # Images — smarter than "any <img> without alt".
        # Matches what Google + WCAG actually care about:
        #   - Missing `alt` attribute entirely = real issue (flag)
        #   - `alt=""` explicitly = decorative, correct pattern (info, not issue)
        #   - 1x1 tracking pixels, aria-hidden / role=presentation, or images
        #     inside an <a>/<button> that already has accessible text → skipped
        #     (the image doesn't need its own alt for screen readers).
        imgs = soup.find_all('img')
        result['images_total'] = len(imgs)
        no_alt_imgs = []              # offending — no alt attr AND not decorative
        no_alt_data = []              # rich per-image records — feeds the missing-alt detail row
        all_images_data = []          # every meaningful image on the page — feeds the All Images panel
        empty_alt_count = 0           # <img alt=""> on genuinely decorative imagery — correct pattern
        empty_alt_content_imgs = []   # <img alt=""> on what looks like content (photo/screenshot/logo) — needs review
        skipped_decorative = 0        # skipped via heuristics (tracker px, aria-hidden, widgets, labeled parent)

        def _has_accessible_name(node):
            """Whether a parent link/button carries its own accessible name —
            text, aria-label, aria-labelledby, or title (own or on a descendant
            img per W3C accessible-name algorithm) — making a missing/empty img
            alt fine for screen readers."""
            if not node:
                return False
            if (node.get('aria-label') or '').strip():
                return True
            if (node.get('aria-labelledby') or '').strip():
                return True
            if (node.get('title') or '').strip():
                return True
            try:
                for d in node.find_all('img'):
                    if (d.get('title') or '').strip():
                        return True
                    if (d.get('aria-label') or '').strip():
                        return True
            except Exception:
                pass
            clone = node
            txt = (clone.get_text(separator=' ', strip=True) or '')
            return len(txt) >= 2

        for img in imgs:
            alt_attr = img.get('alt')  # None if missing, '' if empty, str otherwise
            src = img.get('src', '') or img.get('data-src', '') or ''

            # Decorative / non-content filters — applied to BOTH the
            # missing-alt list and the All Images list. We don't want
            # tracking pixels or hidden spacers polluting either view.
            aria_hidden = (img.get('aria-hidden') or '').lower() == 'true'
            role = (img.get('role') or '').lower()
            w = (img.get('width') or '').strip()
            h = (img.get('height') or '').strip()
            is_hidden = aria_hidden or role in ('presentation', 'none')
            is_tracker = w in ('1', '0') or h in ('1', '0')
            is_data_uri = src.startswith('data:')

            if is_hidden or is_tracker or is_data_uri or not src:
                if alt_attr == '':
                    empty_alt_count += 1
                else:
                    skipped_decorative += 1
                continue

            # Resolve against the post-redirect URL so an http→https
            # redirect doesn't produce phantom http image URLs.
            _img_base = (getattr(resp, 'url', None) or url) if resp is not None else url
            abs_src = urljoin(_img_base, src)

            # Third-party widget filter — reCAPTCHA badges, analytics pixels,
            # chat widgets etc. The site owner can't write alt text for them
            # and they pollute every page that has a contact form.
            if _is_third_party_widget_image(abs_src):
                skipped_decorative += 1
                continue

            # Capture parent + surrounding once per image — used by both
            # the missing-alt detail row and the All Images panel.
            _ptag = img.find_parent(['a', 'button', 'figure'])
            _ptag_name = _ptag.name if _ptag else ''
            _ptext = (_ptag.get_text(separator=' ', strip=True)[:200]
                      if _ptag else '')
            _block = img.find_parent(['p', 'div', 'section', 'article', 'figure', 'li'])
            _surrounding = (_block.get_text(separator=' ', strip=True)[:400]
                            if _block else '')
            # Per W3C accessible-name computation: when alt="" inside a link,
            # the link still has an accessible name if the img carries
            # title/aria-label OR the link itself does. Don't flag those as
            # "empty in link" — screen readers do read them.
            if alt_attr is None:
                _classification = 'missing'
            elif alt_attr.strip() == '':
                in_interactive = _ptag_name in ('a', 'button')
                if in_interactive:
                    img_title = (img.get('title') or '').strip()
                    img_aria = (img.get('aria-label') or '').strip()
                    parent_named = bool(_ptag) and _has_accessible_name(_ptag)
                    if img_title or img_aria or parent_named:
                        _classification = 'empty'
                    else:
                        _classification = 'empty in link'
                elif _filename_looks_decorative(abs_src):
                    _classification = 'empty'
                else:
                    _classification = 'empty (likely content)'
            else:
                _classification = 'present'

            # Record into All Images (every meaningful image, capped at 50/page).
            if len(all_images_data) < 50:
                all_images_data.append({
                    'src': abs_src,
                    'alt': alt_attr,
                    'classification': _classification,
                    'parent_tag': _ptag_name,
                    'parent_text': _ptext,
                    'surrounding': _surrounding,
                })

            # Branch into the existing alt-correctness counters/lists.
            if alt_attr == '':
                if _classification == 'empty (likely content)':
                    empty_alt_content_imgs.append(abs_src)
                    if len(no_alt_data) < 20:
                        no_alt_data.append({
                            'src': abs_src,
                            'alt': alt_attr,
                            'classification': _classification,
                            'parent_tag': _ptag_name,
                            'parent_text': _ptext,
                            'surrounding': _surrounding,
                        })
                else:
                    empty_alt_count += 1
                continue
            if alt_attr is not None and alt_attr.strip():
                continue  # Has meaningful alt — good

            # alt is missing entirely. Skip if the parent link/button
            # already carries the accessible name.
            parent_interactive = img.find_parent(['a', 'button'])
            if parent_interactive and _has_accessible_name(parent_interactive):
                skipped_decorative += 1
                continue

            no_alt_imgs.append(abs_src)
            if len(no_alt_data) < 20:
                no_alt_data.append({
                    'src': abs_src,
                    'alt': alt_attr,
                    'classification': _classification,
                    'parent_tag': _ptag_name,
                    'parent_text': _ptext,
                    'surrounding': _surrounding,
                })

        result['images_no_alt'] = len(no_alt_imgs)
        result['images_no_alt_urls'] = no_alt_imgs[:20]  # cap at 20 per page
        result['images_no_alt_data'] = no_alt_data       # rich per-image records (cap 20)
        result['images_all_data']    = all_images_data   # every meaningful img (cap 50) — feeds All Images panel
        result['images_empty_alt'] = empty_alt_count
        result['images_empty_alt_content'] = len(empty_alt_content_imgs)
        result['images_empty_alt_content_urls'] = empty_alt_content_imgs[:20]
        result['images_decorative_skipped'] = skipped_decorative

        # Links - extract internal + external with anchor text + placement.
        # Placement = which site region the link sits in (nav / header / footer / main),
        # so users can tell boilerplate links from content links — like Screaming Frog.
        def _placement(a_tag):
            for ancestor in a_tag.parents:
                name = (getattr(ancestor, 'name', None) or '').lower()
                if not name: continue
                if name in ('nav',): return 'nav'
                if name in ('header',): return 'header'
                if name in ('footer',): return 'footer'
                if name in ('aside',): return 'sidebar'
                # Check role / class hints
                classes = ' '.join(ancestor.get('class', []) if hasattr(ancestor, 'get') else []).lower()
                role = (ancestor.get('role', '') if hasattr(ancestor, 'get') else '').lower()
                if 'nav' in classes or role == 'navigation': return 'nav'
                if 'footer' in classes: return 'footer'
                if 'header' in classes: return 'header'
                if name == 'main': return 'main'
            return 'body'

        # Resolve relative hrefs against the FINAL URL after redirects, not
        # the request URL. Otherwise crawling http://example.com/ (which 301s
        # to https://example.com/) produces http://example.com/services for
        # every <a href="/services"> in the body — bogus HTTP URLs that don't
        # exist anywhere in the actual HTML.
        link_base = (getattr(resp, 'url', None) or url) if resp is not None else url
        # Honor <base href> when present. The HTML spec says relative URLs
        # resolve against <base href>, not the document URL. Sites that use
        # directory-style URLs (/page/) + relative links without a leading
        # slash (href="other/") otherwise spawn infinite phantom nested paths
        # (/page/other/, /page/other/more/, …) whenever the server returns 200
        # for arbitrary depths — a crawler trap that buries real pages.
        base_tag = soup.find('base', href=True)
        if base_tag:
            base_href = (base_tag.get('href') or '').strip()
            if base_href:
                link_base = urljoin(link_base, base_href)
        int_links = {}      # normalized target -> {anchor, placement}
        ext_links_list = [] # external links captured with anchor + placement
        ext_count = 0
        for a in soup.find_all('a', href=True):
            href = a['href'].strip()
            if not href:
                continue
            href_low = href.lower()
            # Case-insensitive scheme filter (catches Mailto:, MAILTO:, etc.)
            if (href_low.startswith('#')
                or href_low.startswith('javascript:')
                or href_low.startswith('mailto:')
                or href_low.startswith('tel:')
                or href_low.startswith('sms:')
                or href_low.startswith('skype:')
                or href_low.startswith('whatsapp:')
                or href_low.startswith('data:')
                or href_low.startswith('file:')):
                continue
            # Bare email written without the mailto: prefix —
            # <a href="sales@example.com">. urljoin would otherwise produce
            # https://example.com/path/sales@example.com — the email lands
            # in the crawl as a fake URL.
            if '@' in href and _MAILTO_NO_SCHEME_RE.match(href):
                continue
            resolved = urljoin(link_base, href)
            resolved_path_tail = urlparse(resolved).path.rstrip('/').rsplit('/', 1)[-1]
            if '@' in resolved_path_tail and _MAILTO_NO_SCHEME_RE.match(resolved_path_tail):
                continue
            link_domain = urlparse(resolved).netloc.lower().replace('www.', '')
            # Anchor text: prefer visible text, fall back to aria-label / alt of child <img>
            anchor = (a.get_text(separator=' ', strip=True) or '')[:180]
            if not anchor:
                aria = a.get('aria-label') or a.get('title')
                if aria:
                    anchor = aria.strip()[:180]
                else:
                    img_child = a.find('img')
                    if img_child:
                        alt = (img_child.get('alt') or '').strip()[:140]
                        src = (img_child.get('src') or '').strip()
                        # Extract filename so two images with similar alts (e.g. header
                        # logo "Todd Devine Homes text logo" vs tile image "Todd Homes")
                        # can be disambiguated in the inlinks drawer.
                        fname = ''
                        if src:
                            fname = src.split('?', 1)[0].rstrip('/').split('/')[-1][:80]
                        if alt and fname:
                            anchor = f'[image: {alt} — {fname}]'
                        elif alt:
                            anchor = f'[image: {alt}]'
                        elif fname:
                            anchor = f'[image: {fname}]'
                        else:
                            anchor = '[image]'
            if not anchor:
                anchor = '(empty anchor)'
            placement = _placement(a)

            if link_domain == domain:
                normalized = _normalize_crawl_url(resolved)
                if normalized not in int_links:
                    int_links[normalized] = (anchor, placement)
            else:
                ext_count += 1
                if len(ext_links_list) < 300:  # cap payload
                    # Capture rel + target for the External Links report.
                    # rel is multi-token ("nofollow ugc sponsored noopener") —
                    # join with spaces and lowercase for cheap substring checks.
                    rel_attr = a.get('rel') or []
                    if isinstance(rel_attr, list):
                        rel_str = ' '.join(rel_attr).strip().lower()
                    else:
                        rel_str = str(rel_attr).strip().lower()
                    target_attr = (a.get('target') or '').strip().lower()
                    ext_links_list.append([resolved, anchor, placement, rel_str, target_attr])

        # Transport (internal): [[target, anchor, placement], ...]
        # Transport (external): [[target, anchor, placement, rel, target_attr], ...]
        result['internal_link_urls'] = [[t, a, p] for t, (a, p) in int_links.items()]
        result['internal_links'] = len(int_links)
        result['external_link_urls'] = ext_links_list
        result['external_links'] = ext_count

        # Issues detection — skip content/SEO checks for noindex or pagination pages
        # (noindex = Google won't rank it; pagination = archive duplicate, not a canonical page)
        # Also skip URLs that redirected: the resolved target is crawled separately and
        # any content issues belong on that row, not on the 301 source.
        # When ignore_noindex is set, treat noindex pages like indexable ones for
        # the audit so the user sees the full warning/info list, not just the flag.
        # ALSO skip canonicalised pages — they're declared duplicates so any
        # 'Missing meta' / 'Missing title' / 'Thin content' on them is noise;
        # those issues are real on the canonical page and would surface there.
        # The page itself still appears under the 'Canonicalised' report.
        is_canonicalised = result.get('canonical_kind') == 'canonicalised'
        # Status guard: only run content checks on 2xx responses. 3xx/4xx/5xx
        # responses don't have meaningful content; e.g. an unfollowed 301
        # has empty body, so without this guard it would surface "Missing meta
        # description / Missing title / Missing H1" even though the source is
        # just a redirect that never returns HTML.
        _status = result.get('status_code', 0) or 0
        if (result['indexable'] or ignore_noindex) and not result.get('is_pagination') and not result.get('redirect_url') and not is_canonicalised and 200 <= _status < 300:
            if not result['title']:
                result['issues'].append('Missing title')
            elif result['title_len'] > 60:
                result['issues'].append(f'Title too long ({result["title_len"]})')
            elif result['title_len'] < 30:
                result['issues'].append(f'Title too short ({result["title_len"]})')

            if not result['meta_description']:
                result['issues'].append('Missing meta description')
            elif result['meta_len'] > 160:
                result['issues'].append(f'Meta desc too long ({result["meta_len"]})')
            elif result['meta_len'] < 70:
                result['issues'].append(f'Meta desc too short ({result["meta_len"]})')

            # "Missing H1" only fires when no H1 tag has any text — an
            # <h1></h1> wrapping a logo image doesn't make the page
            # "missing" if a populated H1 sits below it.
            h1_count = len(result['h1_list'])
            if not result['h1']:
                result['issues'].append('Missing H1')
            elif h1_count > 1:
                result['issues'].append(f'Multiple H1s ({h1_count})')
            if result['h1'] and result['title'] and result['h1'].strip().lower() == result['title'].strip().lower():
                result['issues'].append('H1 same as title')

            if result['canonical_kind'] == 'missing':
                result['issues'].append('Missing canonical')
            elif result['canonical_kind'] == 'canonicalised':
                result['issues'].append('Canonicalised (points elsewhere)')

            if result['word_count'] < 200:
                result['issues'].append(f'Thin content ({result["word_count"]} words)')

            if result['images_no_alt'] > 0:
                result['issues'].append(f'{result["images_no_alt"]} imgs missing alt')
            if result.get('images_empty_alt_content', 0) > 0:
                result['issues'].append(f'{result["images_empty_alt_content"]} imgs with empty alt on content imagery')

            if not result['schema_types']:
                result['issues'].append('No schema')

            # Viewport
            if not soup.find('meta', attrs={'name': _re.compile(r'^viewport$', _re.I)}):
                result['issues'].append('Missing viewport')

            # Open Graph — any indexable page should have at least og:title + og:image for social sharing
            og = result['og_tags']
            if not og.get('title') and not og.get('image'):
                result['issues'].append('Missing Open Graph tags')
            elif not og.get('image'):
                result['issues'].append('Missing og:image')

            # Twitter Card — not critical but worth flagging
            if not result['twitter_tags']:
                result['issues'].append('Missing Twitter Card')

            # Analytics — flag pages with no tracking at all. But if the site
            # is clearly JS-rendered (Wix, Shopify, Squarespace, Webflow, or
            # client-side React/Vue/Next) and we crawled without JS rendering,
            # the HTML we scanned was the unhydrated shell — so "No analytics
            # detected" is a false negative. Emit a more actionable warning
            # instead so the user knows to re-run with Render JS enabled.
            if not result['analytics']:
                platform = _detect_js_platform(raw_html) if not result.get('js_rendered') else None
                if platform:
                    result['js_platform'] = platform
                    result['issues'].append(f'Analytics unknown — {platform} site needs Render JS')
                else:
                    result['issues'].append('No analytics detected')

        # --- Issues that apply regardless of indexability ---
        # Canonicalised flag — surfaced even when content checks are skipped
        # so the page still appears under the Canonicalised report.
        if is_canonicalised:
            result['issues'].append('Canonicalised (points elsewhere)')
        if result['response_time'] > 3:
            result['issues'].append(f'Slow ({result["response_time"]}s)')

        # URL hygiene (Screaming Frog URL tab)
        for ui in result.get('url_issues', []):
            result['issues'].append(f'URL: {ui}')

        # Mixed content
        if result['mixed_content']:
            result['issues'].append(f'Mixed content ({len(result["mixed_content"])} resources)')

        # Security headers — we still capture their presence in result['security']
        # for the page-detail panel, but we don't flag missing HSTS / X-Content-Type-Options
        # / X-Frame-Options / CSP / Referrer-Policy as issues (low signal-to-noise for SEO).
        sec = result.get('security', {})
        if not sec.get('is_https'):
            result['issues'].append('Served over HTTP (insecure)')

        # JS-vs-non-JS diff. Only when caller opted in AND we successfully
        # rendered a JS version different from the pre-JS HTML. Surfaces
        # JS-only content that's invisible to non-rendering AI crawlers.
        if capture_no_js and result.get('js_rendered') and pre_js_html:
            try:
                base = result.get('url') or url
                nojs = _parse_no_js_subset(pre_js_html, base)
                result['non_js'] = nojs
                result['js_diff'] = _compute_js_diff(result, nojs)
                sev = result['js_diff'].get('severity')
                if sev == 'critical':
                    fields = ', '.join(result['js_diff'].get('fields', [])) or 'title/meta/schema'
                    result['issues'].append(f'JS-only content (critical: {fields})')
                elif sev == 'high':
                    fields = ', '.join(result['js_diff'].get('fields', [])) or 'h1/word_count'
                    result['issues'].append(f'JS-only content (high: {fields})')
            except Exception as e:
                result.setdefault('render_errors', []).append(f'no-js diff: {str(e)[:160]}')

    except requests.exceptions.Timeout:
        result['error'] = 'Timeout'
        result['issues'].append('Timeout')
    except requests.exceptions.SSLError as e:
        # Reached only when even verify=False failed (rare). Still tell the
        # user it's a certificate problem, not a vague "connection error".
        result['error'] = 'SSL certificate error'
        _m = str(e)
        if 'CERTIFICATE_VERIFY_FAILED' in _m or 'unable to get local issuer' in _m or 'self signed' in _m.lower():
            result['issues'].append('SSL certificate verification failed (incomplete chain or untrusted/self-signed cert) — loads in browsers but blocks crawlers')
        else:
            result['issues'].append('SSL error: ' + str(e)[:90])
    except requests.exceptions.ConnectionError as e:
        # Distinguish a true connection failure (DNS/refused/reset) from the
        # generic bucket so the user can act on it.
        result['error'] = 'Connection error'
        _m = str(e).lower()
        if 'name or service not known' in _m or 'failed to resolve' in _m or 'nodename nor servname' in _m:
            result['issues'].append('Connection error — DNS lookup failed (domain not resolving)')
        elif 'refused' in _m:
            result['issues'].append('Connection error — connection refused (server not accepting connections)')
        elif 'reset' in _m:
            result['issues'].append('Connection error — connection reset (often bot/WAF blocking non-browser clients)')
        else:
            result['issues'].append('Connection error')
    except Exception as e:
        result['error'] = str(e)[:100]
        result['issues'].append(f'Error: {str(e)[:60]}')

    return result


def _teardown_pw(pw_page, pw_browser, pw_ctx):
    for name, obj, method in (('page', pw_page, 'close'), ('browser', pw_browser, 'close'), ('pw', pw_ctx, 'stop')):
        if obj is not None:
            try: getattr(obj, method)()
            except Exception: pass



# -------- Sitemap analysis ---------------------------------------------------
# Discovers a site's XML sitemap(s), parses every URL, and diffs them against a
# crawl. Mirrors the Screaming Frog Sitemaps tab. No external API or LLM —
# pure XML parsing + set diffs.
_SITEMAP_DEFAULT_PATHS = (
    '/sitemap.xml', '/sitemap_index.xml', '/sitemap-index.xml',
    '/wp-sitemap.xml',
    '/sitemap.xml.gz',
    '/sitemap1.xml', '/sitemap-1.xml',
    '/post-sitemap.xml', '/page-sitemap.xml',
)
_SITEMAP_NS = '{http://www.sitemaps.org/schemas/sitemap/0.9}'


def _discover_sitemaps(domain):
    """Find sitemap URLs for a domain. Tries robots.txt first, then common
    default paths. When robots.txt points at a sibling subdomain (multisite
    misconfiguration) we surface a warning AND also probe the analysed
    domain's own paths so the diff isn't comparing the crawl against the
    wrong site's URL set.
    """
    found = []
    seen = set()
    warnings = []

    def _add(u, src):
        u = u.strip()
        if u and u not in seen:
            seen.add(u)
            found.append({'url': u, 'source': src})

    analysed_host = (urlparse(domain).netloc or '').lower()

    try:
        r = _http_get(f"{domain.rstrip('/')}/robots.txt", timeout=10,
                         headers={'User-Agent': 'Mozilla/5.0'})
        if r.ok and r.text:
            for line in r.text.splitlines():
                line = line.strip()
                if line.lower().startswith('sitemap:'):
                    sm_url = line.split(':', 1)[1].strip()
                    sm_host = (urlparse(sm_url).netloc or '').lower()
                    src = 'robots.txt'
                    if sm_host and analysed_host and sm_host != analysed_host:
                        src = 'robots.txt (DIFFERENT DOMAIN)'
                        warnings.append(
                            f"robots.txt declares sitemap on a different host ({sm_host}) "
                            f"than the site being analysed ({analysed_host}). "
                            f"Likely multisite misconfiguration — also probing default paths."
                        )
                    _add(sm_url, src)
    except Exception:
        pass

    has_onsite = any((urlparse(s['url']).netloc or '').lower() == analysed_host for s in found)
    if not has_onsite:
        for path in _SITEMAP_DEFAULT_PATHS:
            url = f"{domain.rstrip('/')}{path}"
            try:
                resp = _http_head(url, timeout=8, allow_redirects=True,
                                     headers={'User-Agent': 'Mozilla/5.0'})
                if resp.status_code == 405:
                    resp = _http_get(url, timeout=10, allow_redirects=True,
                                        headers={'User-Agent': 'Mozilla/5.0'},
                                        stream=True)
                    resp.close()
                if resp.ok:
                    ct = (resp.headers.get('content-type') or '').lower()
                    if 'xml' in ct or path.endswith('.xml') or path.endswith('.gz'):
                        _add(url, 'default-path')
                        break
            except Exception:
                pass

    return found, warnings


def _fetch_sitemap_recursive(seed_urls, max_depth=5):
    """Walk a sitemap (handling sitemap-index recursion) and collect every URL."""
    import xml.etree.ElementTree as ET
    import gzip

    urls = []
    sitemaps_meta = []
    errors = []
    visited = set()

    def _walk(sm_url, depth):
        if depth > max_depth or sm_url in visited:
            return
        visited.add(sm_url)
        try:
            r = _http_get(sm_url, timeout=20,
                             headers={'User-Agent': 'Mozilla/5.0',
                                      'Accept': 'application/xml,text/xml,*/*'})
            if not r.ok:
                errors.append({'sitemap': sm_url, 'error': f'http_{r.status_code}'})
                sitemaps_meta.append({'url': sm_url, 'url_count': 0, 'error': f'http_{r.status_code}'})
                return
            content = r.content
            if sm_url.lower().endswith('.gz'):
                try:
                    content = gzip.decompress(content)
                except Exception:
                    pass
            try:
                root = ET.fromstring(content)
            except ET.ParseError as e:
                errors.append({'sitemap': sm_url, 'error': f'xml_parse: {str(e)[:120]}'})
                sitemaps_meta.append({'url': sm_url, 'url_count': 0, 'error': 'xml_parse'})
                return

            tag = root.tag.split('}', 1)[-1] if '}' in root.tag else root.tag
            count_here = 0
            if tag == 'sitemapindex':
                for sm_node in root.findall(f'{_SITEMAP_NS}sitemap'):
                    loc = sm_node.find(f'{_SITEMAP_NS}loc')
                    if loc is not None and loc.text:
                        _walk(loc.text.strip(), depth + 1)
                sitemaps_meta.append({'url': sm_url, 'url_count': 0, 'error': None,
                                      'is_index': True})
            else:
                for url_node in root.findall(f'{_SITEMAP_NS}url'):
                    loc = url_node.find(f'{_SITEMAP_NS}loc')
                    if loc is None or not loc.text:
                        continue
                    lastmod = url_node.find(f'{_SITEMAP_NS}lastmod')
                    urls.append({
                        'url': loc.text.strip(),
                        'lastmod': lastmod.text.strip() if lastmod is not None and lastmod.text else None,
                        'source_sitemap': sm_url,
                    })
                    count_here += 1
                sitemaps_meta.append({'url': sm_url, 'url_count': count_here, 'error': None,
                                      'is_index': False})
        except Exception as e:
            errors.append({'sitemap': sm_url, 'error': str(e)[:200]})
            sitemaps_meta.append({'url': sm_url, 'url_count': 0, 'error': str(e)[:120]})

    for u in seed_urls:
        _walk(u, 0)
    return urls, sitemaps_meta, errors


# File extensions that are NOT HTML pages — sitemaps shouldn't list them
# and the missing-from-sitemap report should skip them entirely.
_NON_HTML_EXTS = (
    '.jpg', '.jpeg', '.png', '.gif', '.webp', '.svg', '.ico', '.bmp', '.tiff', '.avif',
    '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.csv', '.txt', '.rtf',
    '.zip', '.tar', '.gz', '.7z', '.rar',
    '.mp4', '.mov', '.webm', '.m4v', '.avi', '.mkv', '.mp3', '.wav', '.ogg', '.flac',
    '.css', '.js', '.json', '.xml', '.map',
    '.woff', '.woff2', '.ttf', '.otf', '.eot',
)


def _is_non_html_url(u):
    if not u:
        return False
    try:
        path = (urlparse(u).path or '').lower()
    except Exception:
        return False
    return path.endswith(_NON_HTML_EXTS)


def _norm_url(u):
    """Normalise a URL for set-comparison.

    Collapses http vs https, www vs non-www, and trailing slash so the
    sitemap-vs-crawl comparison doesn't flag http variants as "missing from
    sitemap" when the sitemap only lists https URLs. Also strips trailing
    whitespace / %20 — Shopify and other CMSs serve the same page at /foo
    and /foo%20 when an internal <a href> has a trailing space.
    """
    if not u:
        return ''
    u = u.strip()
    try:
        p = urlparse(u)
        host = (p.netloc or '').lower()
        if host.startswith('www.'):
            host = host[4:]
        path = p.path or ''
        while path:
            low = path.lower()
            if low.endswith('%20'):
                path = path[:-3]
            elif path[-1].isspace() or path[-1] == '/':
                path = path[:-1]
            else:
                break
        if not path:
            path = '/'
        return f"https://{host}{path if path != '/' else ''}".lower()
    except Exception:
        return u.rstrip('/').lower()


@app.route('/fetch-robots-txt', methods=['GET'])
def fetch_robots_txt():
    """Fetch a site's robots.txt for the URL filters preview panel."""
    target = (request.args.get('url') or '').strip()
    if not target:
        return jsonify({'error': 'url required'}), 400
    if not target.startswith('http'):
        target = 'https://' + target.lstrip('/')
    try:
        parsed = urlparse(target)
        if not parsed.netloc:
            return jsonify({'error': 'invalid URL'}), 400
        robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
        resp = _http_get(robots_url, timeout=10,
                            headers={'User-Agent': 'Mozilla/5.0 (compatible; OpenSEOCrawler-RobotsPreview)'},
                            allow_redirects=True)
        body = (resp.text or '')[:20000]
        return jsonify({
            'url': robots_url,
            'status': resp.status_code,
            'content': body,
            'length': len(resp.text or ''),
        })
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 200


@app.route('/sitemap-analyse', methods=['POST'])
def _bare_host(url):
    """Normalise a URL/domain to a bare host for matching: strip scheme,
    www., path and trailing slash."""
    h = (url or '').lower().strip()
    h = h.replace('https://', '').replace('http://', '').rstrip('/')
    h = h[4:] if h.startswith('www.') else h
    return h.split('/')[0]


def _same_host_sitemap_urls(sm_urls, domain):
    """Keep only sitemap entries on the same host as the analysed domain.

    robots.txt on multisite setups frequently points at a *sibling* subdomain's
    sitemap (e.g. electricmotors.teco.com.au → appliances.teco.com.au). Those
    URLs belong to a different site and must NOT be diffed against this crawl —
    otherwise every uncrawled foreign URL floods the "in sitemap, not crawled"
    orphan report. Returns (kept_entries, foreign_host_counts) so the caller can
    tell the user exactly what was excluded instead of silently dropping it.
    """
    base = _bare_host(domain)
    if not base:
        return list(sm_urls), {}
    kept, foreign = [], {}
    for entry in sm_urls:
        h = _bare_host(entry.get('url') or '')
        if h and h != base:
            foreign[h] = foreign.get(h, 0) + 1
        else:
            kept.append(entry)
    return kept, foreign


def _foreign_host_warning(foreign_hosts, domain):
    """Human-readable note about sitemap URLs excluded from the diff because
    they live on a different host than the analysed site."""
    if not foreign_hosts:
        return None
    total = sum(foreign_hosts.values())
    detail = ', '.join(f"{h} ({n})" for h, n in sorted(foreign_hosts.items(), key=lambda x: -x[1]))
    return (f"Excluded {total} sitemap URL(s) on a different host — {detail} — from the "
            f"orphan / coverage diff. They belong to another site, not {_bare_host(domain)}, "
            f"so they are not orphans of this crawl.")


def sitemap_analyse():
    """Discover the site's sitemap(s) and diff against a crawl.

    Body: {"domain": "https://example.com", "results": [...page rows...],
           "inlinks": {url: [source_urls...]}}
    """
    data = request.get_json() or {}
    domain = (data.get('domain') or '').rstrip('/')
    if domain and not domain.startswith('http'):
        domain = 'https://' + domain
    results = data.get('results') or []
    inlinks_map = data.get('inlinks') or {}
    manual_sm = (data.get('sitemap_url') or '').strip()

    if not domain:
        return jsonify({'error': 'domain required'}), 400

    if manual_sm:
        if not manual_sm.startswith('http'):
            manual_sm = 'https://' + manual_sm.lstrip('/')
        discovered = [{'url': manual_sm, 'source': 'manual'}]
        discovery_warnings = []
    else:
        discovered, discovery_warnings = _discover_sitemaps(domain)
        if not discovered:
            return jsonify({
                'sitemaps_found': [],
                'warnings': discovery_warnings + ['No sitemap could be discovered. Tried robots.txt and common default paths.'],
                'tried_paths': list(_SITEMAP_DEFAULT_PATHS),
            }), 200

    seed = [d['url'] for d in discovered]
    sm_urls, sitemaps_meta, sm_errors = _fetch_sitemap_recursive(seed)

    # Drop URLs on a different host than the analysed site (multisite robots.txt
    # pointing at a sibling subdomain). They're a different site, not orphans
    # of this crawl.
    sm_urls, _foreign_hosts = _same_host_sitemap_urls(sm_urls, domain)
    _fh_warning = _foreign_host_warning(_foreign_hosts, domain)
    if _fh_warning:
        discovery_warnings = list(discovery_warnings) + [_fh_warning]

    crawl_by_norm = {}
    for r in results:
        u = r.get('url')
        if not u:
            continue
        crawl_by_norm[_norm_url(u)] = r

    sitemap_by_norm = {}
    for entry in sm_urls:
        sitemap_by_norm.setdefault(_norm_url(entry['url']), entry)

    inlinks_by_norm = {}
    for k, v in inlinks_map.items():
        inlinks_by_norm[_norm_url(k)] = v or []

    pag_re = _re.compile(r'/page/\d+/?$|[?&](page|paged|pg)=\d+', _re.I)

    missing_from_sitemap = []
    orphan_in_sitemap = []
    sitemap_only = []
    non_indexable_in_sitemap = []
    non_200_in_sitemap = []
    redirects_in_sitemap = []
    pagination_in_sitemap = []

    for nrm, r in crawl_by_norm.items():
        if nrm in sitemap_by_norm:
            continue
        sc = r.get('status_code') or 0
        if not r.get('indexable', True):
            continue
        if sc and sc != 200:
            continue
        if r.get('redirect_url'):
            continue
        if r.get('is_pagination'):
            continue
        if _is_non_html_url(r.get('url')):
            continue
        # Skip pages whose <link rel=canonical> points elsewhere — they're
        # not the canonical version, so they shouldn't be in the sitemap.
        # Critical on Shopify where /collections/X/products/Y
        # canonicalises to /products/Y.
        canonical = (r.get('canonical') or '').strip()
        if canonical and _norm_url(canonical) != nrm:
            continue
        missing_from_sitemap.append(r.get('url'))

    # Build a canonical→row lookup so a sitemap URL whose canonical version
    # was crawled under a non-canonical alias isn't reported as sitemap-only.
    canonical_to_row = {}
    for r in results:
        can = (r.get('canonical') or '').strip()
        if can:
            canonical_to_row.setdefault(_norm_url(can), r)

    for nrm, entry in sitemap_by_norm.items():
        original_url = entry['url']
        if pag_re.search(urlparse(original_url).path) or pag_re.search('?' + (urlparse(original_url).query or '')):
            pagination_in_sitemap.append(original_url)
        crawled = crawl_by_norm.get(nrm) or canonical_to_row.get(nrm)
        if crawled is None:
            sitemap_only.append({'url': original_url, 'lastmod': entry.get('lastmod')})
            continue
        sc = crawled.get('status_code') or 0
        if sc and sc != 200:
            non_200_in_sitemap.append({'url': original_url, 'status_code': sc})
        # Only flag a "redirect in sitemap" when the redirect lands somewhere
        # OTHER than the sitemap URL. Trailing-slash / case / scheme normalization
        # often makes a crawled URL redirect to the canonical version that the
        # sitemap already lists — that's not a sitemap problem, that's the sitemap
        # being correct. _norm_url strips trailing slashes, so this catches it.
        _rdest = crawled.get('redirect_url')
        if _rdest and _norm_url(_rdest) != nrm:
            redirects_in_sitemap.append({
                'url': original_url,
                'redirects_to': _rdest,
            })
        if not crawled.get('indexable', True):
            non_indexable_in_sitemap.append({
                'url': original_url,
                'reason': 'noindex',
            })
        if (sc == 200 and not crawled.get('redirect_url')
                and crawled.get('indexable', True)
                and not inlinks_by_norm.get(nrm)):
            orphan_in_sitemap.append(original_url)

    warnings = list(discovery_warnings)
    if any(len(s.get('url', '')) and s['url'].startswith('http://') for s in sitemaps_meta):
        warnings.append('At least one sitemap is served over HTTP, not HTTPS.')
    for sm in sitemaps_meta:
        if (sm.get('url_count') or 0) > 50000:
            warnings.append(f"{sm['url']} contains {sm['url_count']} URLs — over the 50,000 sitemap limit.")
    no_lastmod = sum(1 for u in sm_urls if not u.get('lastmod'))
    if sm_urls and no_lastmod / len(sm_urls) > 0.5:
        warnings.append(f"{no_lastmod}/{len(sm_urls)} URLs in sitemap are missing <lastmod>.")

    return jsonify({
        'domain': domain,
        'sitemaps_found': discovered,
        'sitemaps_walked': sitemaps_meta,
        'sitemap_errors': sm_errors,
        'totals': {
            'urls_in_sitemap': len(sm_urls),
            'urls_in_crawl': len(crawl_by_norm),
            'sitemaps_walked': len(sitemaps_meta),
        },
        'reports': {
            'missing_from_sitemap': missing_from_sitemap,
            'orphan_in_sitemap': orphan_in_sitemap,
            'sitemap_only': sitemap_only,
            'non_indexable_in_sitemap': non_indexable_in_sitemap,
            'non_200_in_sitemap': non_200_in_sitemap,
            'redirects_in_sitemap': redirects_in_sitemap,
            'pagination_in_sitemap': pagination_in_sitemap,
        },
        'warnings': warnings,
    })


# =============================================================================
# Near-duplicate content detection (Shingle Jaccard 5-gram + df=1 filter).
# Mirrors the algorithm and API surface in seo-tool. Pure stdlib, no LLM.
# =============================================================================

_ND_STOP = {
    'a', 'an', 'the', 'and', 'or', 'but', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
    'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'should', 'could', 'can',
    'to', 'of', 'in', 'on', 'at', 'by', 'for', 'with', 'about', 'as', 'into', 'through',
    'this', 'that', 'these', 'those', 'i', 'we', 'you', 'they', 'it', 'he', 'she',
    'our', 'your', 'their', 'its', 'his', 'her', 'my',
    'not', 'no', 'so', 'if', 'than', 'then', 'too', 'very', 'just',
    'over', 'under', 'before', 'after', 'between', 'from', 'up', 'down', 'out', 'off',
    'all', 'any', 'each', 'most', 'some', 'other', 'such', 'only', 'own', 'same',
    'us', 'me', 'them', 'who', 'what', 'where', 'when', 'why', 'how',
}

def _nd_tokenize(text):
    return _re.findall(r"[a-z][a-z'\-]{1,}", (text or '').lower())

def _nd_strip_selectors(html_or_text, selectors):
    if not selectors or not html_or_text:
        return html_or_text
    sel_list = [s.strip() for s in selectors.split(',') if s.strip()]
    if not sel_list or '<' not in html_or_text or '>' not in html_or_text:
        return html_or_text
    try:
        soup = BeautifulSoup(html_or_text, 'html.parser')
        for sel in sel_list:
            for node in soup.select(sel):
                node.decompose()
        return soup.get_text(separator=' ', strip=True)
    except Exception:
        return html_or_text

@app.route('/near-dup-content', methods=['POST'])
def near_dup_content():
    payload = request.get_json(silent=True) or {}
    pages = payload.get('pages') or []
    try:
        threshold = float(payload.get('threshold', 0.90))
    except (TypeError, ValueError):
        threshold = 0.90
    threshold = max(0.5, min(0.99, threshold))
    exclude_sel = (payload.get('exclude_selectors') or '').strip()

    t0 = time.perf_counter()

    docs = []
    skipped = 0
    for p in pages:
        url = (p.get('url') or '').strip()
        body = p.get('body_text') or ''
        if not url or not body:
            skipped += 1
            continue
        canonical = (p.get('canonical') or '').strip()
        if canonical and canonical != url:
            skipped += 1
            continue
        if p.get('indexable') is False:
            skipped += 1
            continue
        if exclude_sel:
            body = _nd_strip_selectors(body, exclude_sel)
        toks = [t for t in _nd_tokenize(body) if t not in _ND_STOP and len(t) > 1]
        if len(toks) < 20:
            skipped += 1
            continue
        docs.append({'url': url, 'tokens': toks})

    df = {}
    for d in docs:
        for t in set(d['tokens']):
            df[t] = df.get(t, 0) + 1
    common_terms = {t for t, c in df.items() if c >= 2}

    n = 5
    sets = []
    for d in docs:
        toks = [t for t in d['tokens'] if t in common_terms]
        if len(toks) < n:
            sets.append({'url': d['url'], 'shingles': set()})
            continue
        shingles = {' '.join(toks[i:i + n]) for i in range(len(toks) - n + 1)}
        sets.append({'url': d['url'], 'shingles': shingles})

    pairs = []
    M = len(sets)
    for i in range(M):
        sa = sets[i]['shingles']
        if not sa:
            continue
        for j in range(i + 1, M):
            sb = sets[j]['shingles']
            if not sb:
                continue
            inter = len(sa & sb)
            union = len(sa | sb)
            if not union:
                continue
            sim = inter / union
            if sim >= threshold:
                sample = next(iter(sa & sb), '') if inter else ''
                pairs.append({
                    'url_a': sets[i]['url'],
                    'url_b': sets[j]['url'],
                    'similarity': round(sim, 4),
                    'shared_phrase_sample': sample,
                })
    pairs.sort(key=lambda p: -p['similarity'])

    return jsonify({
        'pairs': pairs,
        'stats': {
            'docs_analysed': M,
            'docs_skipped': skipped,
            'threshold': threshold,
            'took_ms': int((time.perf_counter() - t0) * 1000),
        },
    })


@app.route('/recrawl-url', methods=['POST'])
def recrawl_url():
    """Re-crawl a single URL and return fresh page audit data."""
    import requests as _req
    from urllib.parse import urlparse
    data = request.get_json() or {}
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({'error': 'URL required'}), 400
    domain = urlparse(url).netloc
    try:
        with _req.Session() as session:
            session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,*/*;q=0.9',
                'Accept-Language': 'en-US,en;q=0.9',
            })
            result = _crawl_page(url, session, domain)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/crawl', methods=['POST'])
def crawl_site():
    """BFS site crawl with SSE streaming of per-page results."""
    from collections import deque
    from urllib.parse import urlparse

    data = request.json

    # Resume support: if resume_crawl_id is provided AND we have saved state
    # for it that's not yet expired, restore everything (config + queue +
    # visited + results + inlinks). Otherwise fall through to a fresh crawl.
    resume_id = (data.get('resume_crawl_id') or '').strip()
    resumed_state = None
    if resume_id:
        cached = SUSPENDED_CRAWLS.pop(resume_id, None)
        if cached and (time.time() - cached.get('created', 0)) <= SUSPENDED_CRAWL_TTL:
            resumed_state = cached
        elif cached:
            app.logger.info(f"[crawler] resume {resume_id} expired, falling back to fresh")

    if resumed_state:
        cfg = resumed_state.get('config', {})
        seed_url = resumed_state.get('seed_url') or (data.get('url', '') or '').strip()
        max_pages = int(data.get('max_pages') or cfg.get('max_pages') or 500)
        if max_pages >= 5000:
            max_pages = 999999
        max_depth = min(int(cfg.get('max_depth', 10) or 10), 20)
        crawl_delay = max(float(cfg.get('crawl_delay', 0.4) or 0.4), 0.0)
        render_js = bool(cfg.get('render_js', False))
        ignore_robots = bool(cfg.get('ignore_robots', False))
        ignore_noindex = bool(cfg.get('ignore_noindex', False))
        compare_no_js = bool(cfg.get('compare_no_js', False)) and render_js
    else:
        seed_url = (data.get('url', '') or '').strip()
        if not seed_url:
            return json.dumps({'error': 'URL is required'}), 400
        if not seed_url.startswith('http'):
            seed_url = 'https://' + seed_url

        max_pages = int(data.get('max_pages', 500) or 500)
        if max_pages >= 5000:
            max_pages = 999999  # unlimited
        max_depth = min(int(data.get('max_depth', 10) or 10), 20)
        crawl_delay = max(float(data.get('crawl_delay', 0.4) or 0.4), 0.0)
        render_js = bool(data.get('render_js', False))
        ignore_robots = bool(data.get('ignore_robots', False))
        ignore_noindex = bool(data.get('ignore_noindex', False))
        # JS vs non-JS compare. Gated on render_js — only meaningful when
        # we have something to compare against.
        compare_no_js = bool(data.get('compare_no_js', False)) and render_js
    # Concurrent workers. Default 5 matches Screaming Frog. Clamped to [1, 20].
    # When render_js is on, Playwright can't share a single page across threads —
    # force single-worker mode so page state stays consistent.
    max_workers = int(data.get('max_workers', 5) or 5)
    max_workers = max(1, min(20, max_workers))
    if render_js:
        max_workers = 1

    # URL include/exclude patterns — robots.txt syntax (Google's spec):
    #   *       matches any sequence
    #   $       at end anchors end of URL
    #   ?, .    are LITERAL (no fnmatch single-char wildcard surprise)
    # Exclude beats include. If include is non-empty, URLs must match at least one.
    # Stored in ACTIVE_CRAWL_RULES so /crawl/update-rules can mutate them mid-crawl.
    import uuid as _uuid
    def _parse_patterns(raw):
        if not raw: return []
        return [p.strip() for p in raw.splitlines() if p.strip() and not p.strip().startswith('#')]
    crawl_id = _uuid.uuid4().hex[:12]
    # crawl_delay also lives here so /crawl/update-rules can mutate it mid-crawl
    # (the slider in the UI auto-pushes the new value while a crawl is running).
    ACTIVE_CRAWL_RULES[crawl_id] = {
        'include': _parse_patterns(data.get('include_patterns', '')),
        'exclude': _parse_patterns(data.get('exclude_patterns', '')),
        'crawl_delay': crawl_delay,
    }
    ACTIVE_CRAWL_LIMITS[crawl_id] = {
        'max_pages': max_pages,
        'continue_event': threading.Event(),
        'finalize': False,
        'bumps': 0,
    }

    def _current_max():
        return (ACTIVE_CRAWL_LIMITS.get(crawl_id) or {}).get('max_pages', max_pages)

    # /cdn-cgi/ — Cloudflare infra paths injected by the proxy. The most
    # common is /cdn-cgi/l/email-protection (the obfuscated-email endpoint
    # Cloudflare auto-injects) which returns 404 when fetched directly
    # because it's only meant to be loaded as a script via the email
    # obfuscation. /cdn-cgi/scripts/, /cdn-cgi/challenge-platform/,
    # /cdn-cgi/rum, /cdn-cgi/bm/ etc. are all infra, not site content.
    # Skipping the whole prefix keeps the 404 report focused on real pages.
    _NON_PAGE_PATH_FRAGMENTS = (
        '/feed/', '/feed.atom', '/feed.rss', '/comments/feed/',
        '/wp-json/', '/wp-admin/', '/wp-login.php', '/xmlrpc.php',
        '/?wc-ajax=', '/cart/?', '/checkout/?',
        '/sitemap.xml', '/sitemap_index.xml',
        '/cdn-cgi/',
    )

    def _url_allowed(u):
        # Drop media files + non-HTML endpoints so they never enter the
        # results table or pollute per-page reports (Schema by Page, etc.).
        if _is_non_html_url(u):
            return False
        try:
            _pu = urlparse(u)
            path_lower = (_pu.path or '').lower()
            query_lower = (_pu.query or '').lower()
        except Exception:
            path_lower = ''
            query_lower = ''
        if any(frag in path_lower for frag in _NON_PAGE_PATH_FRAGMENTS):
            return False
        # Page-builder AJAX pagination/filter traps. Elementor Pro's Posts/Loop
        # widget paginates via ?e-page-<widgetid>=N (and filters via
        # ?e-filter-<id>=...), so a single listing page exposes hundreds of
        # query-string variants — /commentary/ on a real site linked
        # ?e-page-...=854 — each a rel=canonical duplicate of the base page.
        # A link-following crawler chases every one, AND each variant re-exposes
        # the next, so the queue explodes combinatorially toward infinity (the
        # classic "26k queued on a 9k-page site" blow-up). These are never
        # content you want indexed, so we skip them by default — matched on the
        # query-param KEY so it's widget-id- and page-number-agnostic. Jetpack
        # Infinite Scroll (?infinity) gets the same treatment.
        if query_lower:
            for _part in query_lower.split('&'):
                _qk = _part.split('=', 1)[0].strip()
                if _qk.startswith('e-page-') or _qk.startswith('e-filter-') or _qk == 'infinity':
                    return False
        # Repeating-segment guard: refuses URLs where the same path slug
        # appears more than once (e.g. /our-team/our-team/ or
        # /case-studies/our-team/case-studies/). This is the classic
        # crawler trap caused by sites whose nav uses path-relative hrefs
        # ("our-team/" instead of "/our-team/") - resolved against any
        # sub-page, the relative link bleeds the parent path infinitely.
        # Legitimate URLs almost never repeat the same slug 2+ times in a
        # path, so this is a high-precision filter. Short numeric segments
        # ("/2026/01/01/") are excluded from the dedup check so date-based
        # archives still resolve.
        try:
            _segs = [s for s in path_lower.strip('/').split('/') if s]
            _word_segs = [s for s in _segs if not s.isdigit() and len(s) >= 4]
            if len(_word_segs) != len(set(_word_segs)):
                return False
        except Exception:
            pass
        rules = ACTIVE_CRAWL_RULES.get(crawl_id) or {}
        excl = rules.get('exclude') or []
        incl = rules.get('include') or []
        if excl and any(_robots_pattern_match(p, u) for p in excl):
            return False
        if incl and not any(_robots_pattern_match(p, u) for p in incl):
            return False
        return True

    parsed = urlparse(seed_url)
    domain = parsed.netloc.lower().replace('www.', '')

    app.logger.info(f"[crawler] Starting crawl of {seed_url} (max={max_pages}, depth={max_depth}, delay={crawl_delay}s, js={render_js}) from {request.remote_addr}")

    def generate():
        # Fetch robots.txt up-front so the user sees what we're following.
        # If ignore_robots is set, we still fetch (informational) but won't enforce.
        # We replace stdlib urllib.robotparser with our own wildcard-aware
        # checker because stdlib silently drops '*' and '$' patterns - any
        # site whose robots.txt uses 'Disallow: *?swoof*' or 'Disallow: *.pdf*'
        # would otherwise leak those URLs into the crawl.
        robots_can_fetch = lambda u: True
        robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
        robots_status = 'fetching'
        robots_rules = 0
        robots_issues = []        # AI/search-engine blocking flags, attached to the homepage row
        robots_attached = False   # attach once, to the first (seed) page emitted
        try:
            resp = _http_get(robots_url, timeout=8, headers={'User-Agent': 'SEO-Audit-Bot'})
            if resp.status_code == 200:
                robots_can_fetch = _build_robots_checker(resp.text)
                robots_rules = resp.text.count('Disallow')
                robots_status = 'ignored' if ignore_robots else 'respecting'
                robots_issues = _analyze_robots_txt(resp.text)
                yield f"data: {json.dumps({'type':'info','msg':f'Downloaded robots.txt ({robots_rules} Disallow rules) — {robots_status}'})}\n\n"
                if robots_issues:
                    yield f"data: {json.dumps({'type':'info','msg':'robots.txt: ' + '; '.join(robots_issues)})}\n\n"
            else:
                robots_status = 'not found'
                yield f"data: {json.dumps({'type':'info','msg':f'robots.txt returned HTTP {resp.status_code} — no rules to enforce'})}\n\n"
        except Exception as e:
            robots_status = 'error'
            yield f"data: {json.dumps({'type':'info','msg':f'robots.txt unreachable ({str(e)[:80]}) — continuing without'})}\n\n"

        session = requests.Session()
        # Realistic Chrome header set. Default python-requests UA + empty
        # Accept-Language triggers 403 on Shopify/Cloudflare-fronted stores.
        # Sec-Fetch-Site=same-origin is correct for crawl traffic (we follow
        # links from a previously fetched page on the same host).
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-AU,en;q=0.9,en-US;q=0.8',
            'Accept-Encoding': 'gzip, deflate',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Ch-Ua': '"Chromium";v="132", "Not_A Brand";v="24"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-User': '?1',
        })

        # Launch Playwright browser once per crawl if JS rendering requested
        pw_ctx = None
        pw_browser = None
        pw_page = None
        if render_js:
            try:
                from playwright.sync_api import sync_playwright
                pw_ctx = sync_playwright().start()
                pw_browser = pw_ctx.chromium.launch(headless=True, args=['--no-sandbox', '--disable-dev-shm-usage'])
                pw_page = pw_browser.new_page(
                    viewport={'width': 1280, 'height': 900},
                    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36'
                )
                pw_page.set_default_timeout(20000)
                yield f"data: {json.dumps({'type': 'info', 'msg': 'JS rendering enabled (Playwright). Crawl will be 3-5x slower.'})}\n\n"
            except Exception as e:
                app.logger.warning(f"[crawler] Playwright init failed: {e}")
                yield f"data: {json.dumps({'type': 'info', 'msg': f'JS rendering unavailable ({str(e)[:100]}); using raw HTML only.'})}\n\n"
                pw_page = None

        if resumed_state:
            queue = deque(tuple(item) for item in resumed_state.get('queue', []))
            visited = set(resumed_state.get('visited', []))
            results = list(resumed_state.get('results', []))
            errors = int(resumed_state.get('errors', 0))
            total_time = float(resumed_state.get('total_time', 0))
            inlinks_map = dict(resumed_state.get('inlinks_map', {}))
            yield f"data: {json.dumps({'type': 'resumed', 'previous_pages': len(results), 'queued': len(queue), 'errors': errors})}\n\n"
        else:
            queue = deque()
            _seed_norm = _normalize_crawl_url(seed_url)
            queue.append((_seed_norm, 0))
            # Pre-mark the seed (and its slash-alt) as visited so any page that
            # links back to the homepage doesn't re-queue it. visited =
            # "URL has been seen / queued / processed".
            visited = {_seed_norm}
            _seed_alt = _crawl_slash_alt(_seed_norm)
            if _seed_alt:
                visited.add(_seed_alt)
            results = []
            errors = 0
            total_time = 0
            # Map of target URL -> list of source URLs linking to it (inlinks / Screaming Frog style)
            inlinks_map = {}

        tracked_kws = []  # left in place for payload compatibility; no external data sources in the public build
        yield f"data: {json.dumps({'type': 'start', 'domain': domain, 'workers': max_workers, 'crawl_id': crawl_id})}\n\n"

        # CMS fingerprint using the seed page (cheap HEAD+GET already handles this below,
        # but we want the info up-front so the UI can badge + offer one-click recommendations).
        try:
            cms_resp = session.get(seed_url, timeout=10, allow_redirects=True)
            cms_info = detect_cms(seed_url, cms_resp.text, dict(cms_resp.headers))
            if cms_info.get('cms'):
                prof = CMS_PROFILES.get(cms_info['cms'], {})
                cms_info['profile'] = {
                    'exclude_patterns': prof.get('exclude_patterns', []),
                    'suggested_settings': prof.get('suggested_settings', {}),
                    'schema_warnings': prof.get('schema_warnings', []),
                    'tips': prof.get('tips', []),
                }
            yield f"data: {json.dumps({'type': 'cms_detected', **cms_info})}\n\n"
        except Exception as _cms_err:
            app.logger.info(f"[crawler] CMS detection skipped: {_cms_err}")

        # Per-host politeness: minimum gap between two requests to the same host.
        # Workers on DIFFERENT hosts run freely; same-host workers serialise via this lock+timestamp map.
        from concurrent.futures import ThreadPoolExecutor, wait as _fwait, FIRST_COMPLETED
        from urllib.parse import urlparse as _up
        host_last_fetch = {}
        host_lock = threading.Lock()
        host_backoff = {}      # host → multiplier for adaptive slow-down on 429/403/503
        host_pause_until = {}  # host → epoch-seconds to halt ALL workers for this host

        def _wait_host_turn(u):
            # Reads crawl_delay live from ACTIVE_CRAWL_RULES so the slider in
            # the UI can change politeness mid-crawl. Falls back to the local
            # if the rules entry has been cleared (post-stop / cleanup).
            # Honours BOTH the normal per-host delay AND a hard pause window
            # set after a 429/403/503. The pause blocks every worker targeting
            # this host, not just the one that saw the error.
            host = _up(u).netloc
            while True:
                _live = (ACTIVE_CRAWL_RULES.get(crawl_id) or {}).get('crawl_delay')
                live_delay = float(_live) if _live is not None else crawl_delay
                with host_lock:
                    now = time.time()
                    last = host_last_fetch.get(host, 0)
                    pause_until = host_pause_until.get(host, 0)
                    backoff = host_backoff.get(host, 1.0)
                    effective_delay = live_delay * backoff
                    wait = max((last + effective_delay) - now, pause_until - now)
                    if wait <= 0:
                        host_last_fetch[host] = now
                        return
                time.sleep(min(wait, 5))  # wake periodically in case pause is extended

        def _adjust_host_backoff(u, page_data):
            host = _up(u).netloc
            status = page_data.get('status_code', 0)
            waf = page_data.get('_waf_block')  # 'wordfence'/'cloudflare'/'sucuri'/'siteground' or None
            with host_lock:
                cur = host_backoff.get(host, 1.0)
                if status in (429, 503):
                    # Hard slow-down on the first hit: jump straight to 10×.
                    host_backoff[host] = max(cur * 2, 10.0) if cur < 10.0 else min(cur * 2, 40.0)
                    hint = float(page_data.get('_retry_hint') or 0)
                    existing = host_pause_until.get(host, 0) - time.time()
                    if waf:
                        # WAF tripped — 30s won't help (Wordfence default block is 5-60min).
                        # Pause 5 min minimum and jump backoff to 20× so any resume crawls
                        # at a crawl.
                        pause_secs = max(hint, 300.0, existing)
                        host_backoff[host] = max(cur * 5, 20.0) if cur < 20.0 else min(cur * 2, 40.0)
                    else:
                        # Plain transient 503 (host briefly down, not WAF) — short pause.
                        pause_secs = max(hint, 30.0, existing)
                    host_pause_until[host] = time.time() + pause_secs
                elif status == 403:
                    # 403 = Cloudflare/Shopify bot block. Treat like 429 — pause
                    # the host so all workers stop hammering, and escalate backoff
                    # hard so when we resume we're an order of magnitude slower.
                    host_backoff[host] = max(cur * 3, 10.0) if cur < 10.0 else min(cur * 2, 30.0)
                    existing = host_pause_until.get(host, 0) - time.time()
                    if waf:
                        pause_secs = max(300.0, existing)
                        host_backoff[host] = max(cur * 5, 20.0) if cur < 20.0 else min(cur * 2, 40.0)
                    else:
                        pause_secs = max(60.0, existing)
                    host_pause_until[host] = time.time() + pause_secs
                elif status == 0 or page_data.get('error'):
                    # Transient connection/network error — bump backoff lightly,
                    # don't pause the host (other URLs may still work fine).
                    host_backoff[host] = min(cur * 1.5, 20.0)
                else:
                    # Decay back toward 1.0 on successes
                    if cur > 1.0:
                        host_backoff[host] = max(1.0, cur * 0.8)

        def _fetch_job(url, depth):
            """Worker: politeness-wait, fetch, return (url, depth, page_data)."""
            _wait_host_turn(url)
            pd = _crawl_page(url, session, domain, pw_page=pw_page, ignore_noindex=ignore_noindex, capture_no_js=compare_no_js)
            pd['depth'] = depth
            _adjust_host_backoff(url, pd)
            return url, depth, pd

        in_flight = {}  # future -> (url, depth)
        consecutive_errors = 0

        def _visit(url):
            visited.add(url)
            alt = _crawl_slash_alt(url)
            if alt:
                visited.add(alt)

        def _dequeue_next():
            """Pop the next URL that passes filters + robots. Returns (url, depth) or None.
            URLs are added to `visited` at enqueue time now (to prevent the same URL
            from being queued N times when N pages link to it), so we don't gate on
            visited here — it would skip every URL since they're all in visited."""
            while queue:
                url, depth = queue.popleft()
                if depth > max_depth:
                    continue
                if not _url_allowed(url):
                    continue
                if not ignore_robots:
                    try:
                        if not robots_can_fetch(url):
                            continue
                    except Exception:
                        pass
                return url, depth
            return None

        try:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
              # Outer loop lets /crawl/continue bump the page cap and resume
              # without restarting the crawl from scratch.
              while True:
                # Prime the pool
                while len(in_flight) < max_workers and (len(results) + len(in_flight)) < _current_max():
                    nxt = _dequeue_next()
                    if nxt is None:
                        break
                    fut = executor.submit(_fetch_job, nxt[0], nxt[1])
                    in_flight[fut] = nxt

                while in_flight:
                    done_set, _ = _fwait(in_flight.keys(), return_when=FIRST_COMPLETED)
                    for fut in done_set:
                        submitted = in_flight.pop(fut, None)
                        try:
                            url, depth, page_data = fut.result()
                        except Exception as e:
                            url = submitted[0] if submitted else '?'
                            page_data = {
                                'url': url, 'status_code': 0, 'error': str(e)[:200],
                                'issues': [f'Fetch error: {str(e)[:80]}'], 'depth': submitted[1] if submitted else 0,
                            }

                        # Count adaptive signal for surfacing a speed notice
                        status = page_data.get('status_code', 0)
                        _waf_kind = page_data.get('_waf_block')
                        if _waf_kind:
                            # WAF block — surface it immediately, regardless of streak count.
                            # 5-min pause is already applied via _adjust_host_backoff.
                            _pause = max(0, host_pause_until.get(_up(url).netloc, 0) - time.time())
                            yield f"data: {json.dumps({'type': 'speed_adjusted', 'reason': f'{_waf_kind.title()} blocked us — host paused {int(_pause)}s, backoff escalated'})}\n\n"
                        if status in (429, 503, 403) or status == 0 or page_data.get('error'):
                            consecutive_errors += 1
                            if consecutive_errors in (3, 6, 12):
                                _reason_code = status if status else 'conn'
                                yield f"data: {json.dumps({'type': 'speed_adjusted', 'reason': f'HTTP {_reason_code} — per-host back-off active'})}\n\n"
                        else:
                            consecutive_errors = max(0, consecutive_errors - 1)

                        link_urls = page_data.get('internal_link_urls', [])

                        # Tracked-keyword enrichment
                        if tracked_kws:
                            page_url_clean = url.rstrip('/')
                            matching_kws = [k for k in tracked_kws if k.get('url','').rstrip('/') == page_url_clean]
                            if matching_kws:
                                page_data['tracked_keywords'] = matching_kws

                        # If a redirect happened, _crawl_page rewrote page_data['url']
                        # to the canonical final URL. Mark BOTH the requested URL and
                        # the canonical URL as visited so a later internal link to
                        # the canonical URL doesn't queue a duplicate row.
                        canonical_url = page_data.get('url')
                        if canonical_url and canonical_url != url:
                            visited.add(canonical_url)
                            _alt = _crawl_slash_alt(canonical_url)
                            if _alt:
                                visited.add(_alt)
                            if any(r.get('url') == canonical_url for r in results):
                                continue

                        results.append(page_data)
                        if page_data.get('error'):
                            errors += 1
                        total_time += page_data.get('response_time', 0)

                        # Enqueue discovered links + record inlinks
                        source_url = page_data.get('url') or url
                        for entry in link_urls:
                            if isinstance(entry, (list, tuple)):
                                link = entry[0]
                                anchor = entry[1] if len(entry) > 1 else ''
                                placement = entry[2] if len(entry) > 2 else ''
                            else:
                                link, anchor, placement = entry, '', ''
                            bucket = inlinks_map.setdefault(link, [])
                            key = (source_url, anchor, placement)
                            if not any((e.get('source'), e.get('anchor'), e.get('placement')) == key for e in bucket):
                                bucket.append({'source': source_url, 'anchor': anchor, 'placement': placement})
                            alt = _crawl_slash_alt(link)
                            if link not in visited and (not alt or alt not in visited) and _url_allowed(link):
                                # Mark on enqueue (not dequeue) so the same URL
                                # discovered from N pages doesn't get queued N
                                # times before it's pulled.
                                visited.add(link)
                                if alt:
                                    visited.add(alt)
                                queue.append((link, depth + 1))

                        # Site-level robots.txt findings ride on the homepage row
                        # (first non-error page emitted) so they show up in the
                        # issues list/sidebar like every other red issue.
                        if robots_issues and not robots_attached and not page_data.get('error'):
                            page_data['issues'] = list(robots_issues) + (page_data.get('issues') or [])
                            robots_attached = True

                        from itertools import islice as _islice
                        queue_sample = [u for u, _d in _islice(queue, 100)]
                        yield f"data: {json.dumps({'type': 'page', 'data': page_data, 'crawled': len(results), 'queued': len(queue), 'errors': errors, 'queue_sample': queue_sample})}\n\n"

                    # Keep the pool topped up
                    while len(in_flight) < max_workers and (len(results) + len(in_flight)) < _current_max():
                        nxt = _dequeue_next()
                        if nxt is None:
                            break
                        fut2 = executor.submit(_fetch_job, nxt[0], nxt[1])
                        in_flight[fut2] = nxt

                # In-flight drained. Either queue is exhausted, the user asked
                # us to finalize, or we hit the page cap with URLs still queued.
                _state = ACTIVE_CRAWL_LIMITS.get(crawl_id) or {}
                if _state.get('finalize') or not queue:
                    break

                # Hit the cap with URLs still queued — surface a prompt and
                # wait for the user to either bump the cap or finalize.
                from itertools import islice as _islice
                queue_sample = [u for u, _d in _islice(queue, 100)]
                yield f"data: {json.dumps({'type': 'limit_reached', 'queued': len(queue), 'fetched': len(results), 'current_max': _current_max(), 'crawl_id': crawl_id, 'queue_sample': queue_sample})}\n\n"
                _ev = _state.get('continue_event')
                if _ev is None:
                    break
                while not _ev.wait(timeout=15):
                    yield ': waiting\n\n'  # SSE comment heartbeat
                _ev.clear()
                _state2 = ACTIVE_CRAWL_LIMITS.get(crawl_id) or {}
                if _state2.get('finalize'):
                    break
                yield f"data: {json.dumps({'type': 'crawl_resumed', 'new_max': _current_max(), 'queued': len(queue)})}\n\n"

        except GeneratorExit:
            # Re-enqueue any in-flight URLs so resume picks them up. URL stays
            # in `visited` (it's been seen) — dequeue doesn't gate on visited.
            for _fut, _submitted in list(in_flight.items()):
                if not _submitted:
                    continue
                _u, _d = _submitted
                queue.appendleft((_u, _d))
            # Prune anything older than TTL while we're here
            _now = time.time()
            for _cid in list(SUSPENDED_CRAWLS.keys()):
                if _now - SUSPENDED_CRAWLS[_cid].get('created', 0) > SUSPENDED_CRAWL_TTL:
                    SUSPENDED_CRAWLS.pop(_cid, None)
            # Persist state for resume
            SUSPENDED_CRAWLS[crawl_id] = {
                'created': _now,
                'seed_url': seed_url,
                'domain': domain,
                'queue': list(queue),
                'visited': list(visited),
                'results': results,
                'inlinks_map': inlinks_map,
                'errors': errors,
                'total_time': total_time,
                'config': {
                    'max_pages': max_pages,
                    'max_depth': max_depth,
                    'crawl_delay': crawl_delay,
                    'render_js': render_js,
                    'compare_no_js': compare_no_js,
                    'ignore_robots': ignore_robots,
                    'ignore_noindex': ignore_noindex,
                    'max_workers': max_workers,
                },
            }
            app.logger.info(f"[crawler] {crawl_id} suspended (resumable for {SUSPENDED_CRAWL_TTL//60}m): {len(results)} done, {len(queue)} queued")
            session.close()
            _teardown_pw(pw_page, pw_browser, pw_ctx)
            ACTIVE_CRAWL_RULES.pop(crawl_id, None)
            ACTIVE_CRAWL_LIMITS.pop(crawl_id, None)
            return

        session.close()
        _teardown_pw(pw_page, pw_browser, pw_ctx)

        # Summary
        avg_time = round(total_time / len(results), 2) if results else 0
        issue_counts = {}
        for r in results:
            for issue in r.get('issues', []):
                # Normalize issue name for counting
                base = issue.split('(')[0].strip()
                issue_counts[base] = issue_counts.get(base, 0) + 1

        summary = {
            'total': len(results),
            'errors': errors,
            'warnings': sum(1 for r in results if r.get('issues')),
            'avg_time': avg_time,
            'issue_counts': issue_counts,
            'js_rendered_count': sum(1 for r in results if r.get('js_rendered')),
            'render_js': render_js,
            'compare_no_js': compare_no_js,
            'js_diff_counts': {
                'critical': sum(1 for r in results if (r.get('js_diff') or {}).get('severity') == 'critical'),
                'high':     sum(1 for r in results if (r.get('js_diff') or {}).get('severity') == 'high'),
                'medium':   sum(1 for r in results if (r.get('js_diff') or {}).get('severity') == 'medium'),
                'none':     sum(1 for r in results if (r.get('js_diff') or {}).get('severity') == 'none'),
                'pages_with_diff': sum(1 for r in results if r.get('js_diff') and (r.get('js_diff') or {}).get('severity') != 'none'),
                'total_compared': sum(1 for r in results if r.get('js_diff')),
            },
        }

        # Attach inlinks per page (cap to 20 for payload size)
        inlinks_payload = {}
        for r in results:
            u = r.get('url')
            if not u:
                continue
            # A redirected row's url is the FINAL destination, but inbound links
            # were recorded against the originally-linked (redirecting) URL.
            # Emit under both keys so the Redirects view can show "pages linking
            # to the redirecting URL — update these" instead of an empty list.
            keys = [u]
            ou = r.get('original_url')
            if ou and ou != u:
                keys.append(ou)
            for k in keys:
                # Look up by exact URL and normalized variants
                sources = inlinks_map.get(k) or inlinks_map.get(k.rstrip('/')) or inlinks_map.get(k + '/') or []
                if sources:
                    inlinks_payload[k] = sources[:20]

        # ---------- Post-crawl aggregated reports ----------
        from collections import defaultdict as _dd

        # Duplicate titles / metas / H1s / body
        # Skip redirected URLs (http/https/www variants that 301 to the canonical)
        # and non-200 responses — those aren't unique content, just transit stops.
        # Dedupe within each group by normalised URL so pagination (/page/2/)
        # and tracking/ecommerce params (?add-to-cart=, ?utm_*, ?replytocom=)
        # don't fragment a single canonical page across its variants.
        def _group_by(field_getter):
            g = _dd(dict)  # value -> {normalised_url: original_url}
            for r in results:
                if not (r.get('indexable', True) or ignore_noindex):
                    continue
                if r.get('redirect_url'):
                    continue
                if r.get('status_code') and r['status_code'] >= 300:
                    continue
                # Skip canonicalised-elsewhere pages — rel=canonical already
                # declares them duplicates of another URL, so flagging here
                # is double-counting. Critical on Shopify where the same
                # product is reachable via /products/X and
                # /collections/Y/products/X (canonical points to bare URL).
                url = r.get('url') or ''
                canonical = (r.get('canonical') or '').strip()
                if canonical and _norm_url(canonical) != _norm_url(url):
                    continue
                v = field_getter(r)
                if not v:
                    continue
                norm = _normalize_url_for_dup(url)
                if norm not in g[v]:
                    g[v][norm] = url
            return {k: list(d.values()) for k, d in g.items() if len(d) > 1}

        dup_titles = _group_by(lambda r: (r.get('title') or '').strip().lower())
        dup_metas = _group_by(lambda r: (r.get('meta_description') or '').strip().lower())
        dup_h1s = _group_by(lambda r: (r.get('h1') or '').strip().lower())
        dup_bodies = _group_by(lambda r: r.get('body_hash') or '')

        # Response code summary
        rc_buckets = {'2xx': 0, '3xx': 0, '4xx': 0, '5xx': 0, 'other': 0}
        for r in results:
            sc = r.get('status_code', 0)
            if 200 <= sc < 300: rc_buckets['2xx'] += 1
            elif 300 <= sc < 400: rc_buckets['3xx'] += 1
            elif 400 <= sc < 500: rc_buckets['4xx'] += 1
            elif 500 <= sc < 600: rc_buckets['5xx'] += 1
            else: rc_buckets['other'] += 1

        # Redirect chains (2+ hops) — surfaced distinctly from single redirects
        redirect_chains = [
            {'url': r['url'], 'chain': r['redirect_chain'], 'hops': r['redirect_hops']}
            for r in results if r.get('redirect_hops', 0) >= 2
        ]

        # Orphan detection via sitemap (URLs in sitemap but not in crawl).
        # Handle sitemap index by recursively fetching sub-sitemaps (max 20 to bound cost).
        orphans = []
        sitemap_urls_set = set()
        import xml.etree.ElementTree as _ET

        def _fetch_sitemap(sm_url, depth=0, seen=None):
            if seen is None: seen = set()
            if depth > 2 or sm_url in seen or len(seen) > 20:
                return
            seen.add(sm_url)
            try:
                r = _http_get(sm_url, timeout=10, headers={'User-Agent': 'SEO-Audit-Bot'})
                if r.status_code != 200:
                    return
                root = _ET.fromstring(r.content)
                tag = root.tag.lower()
                if tag.endswith('sitemapindex'):
                    # Recurse into each sub-sitemap
                    for loc in root.findall('.//{http://www.sitemaps.org/schemas/sitemap/0.9}loc'):
                        if loc.text:
                            _fetch_sitemap(loc.text.strip(), depth + 1, seen)
                else:
                    # urlset — collect page URLs
                    for loc in root.findall('.//{http://www.sitemaps.org/schemas/sitemap/0.9}loc'):
                        if loc.text:
                            sitemap_urls_set.add(loc.text.strip().rstrip('/'))
            except Exception:
                pass

        _fetch_sitemap(f"{parsed.scheme}://{parsed.netloc}/sitemap.xml")
        crawled_urls_set = {r['url'].rstrip('/') for r in results}
        # Also include redirect targets (since they were actually reached)
        for r in results:
            if r.get('redirect_url'):
                crawled_urls_set.add(r['redirect_url'].rstrip('/'))
        orphans = sorted(sitemap_urls_set - crawled_urls_set)[:200]

        # Crawl depth distribution (depth tracking needs to be added in BFS queue,
        # stored on page_data via tuple queue; for now derive via shortest inlink path
        # approximation — homepage=0, pages linked from homepage=1, etc.)
        # Simpler: use what we have — each page's 'depth' is set by BFS via queue tuple.
        depth_dist = _dd(int)
        for r in results:
            depth_dist[r.get('depth', 0)] += 1

        reports = {
            'response_codes': rc_buckets,
            'duplicate_titles': [{'value': k, 'urls': v} for k, v in sorted(dup_titles.items(), key=lambda x: -len(x[1]))][:100],
            'duplicate_metas': [{'value': k, 'urls': v} for k, v in sorted(dup_metas.items(), key=lambda x: -len(x[1]))][:100],
            'duplicate_h1s': [{'value': k, 'urls': v} for k, v in sorted(dup_h1s.items(), key=lambda x: -len(x[1]))][:100],
            'duplicate_bodies': [{'value': k[:8], 'urls': v} for k, v in sorted(dup_bodies.items(), key=lambda x: -len(x[1]))][:100],
            'redirect_chains': redirect_chains[:200],
            'orphans': orphans,
            'sitemap_count': len(sitemap_urls_set),
            'depth_distribution': dict(depth_dist),
        }

        # If the crawl ended with just the start page (or nothing), tell the
        # user WHY instead of silently stopping. Most reports of "it only
        # crawled the homepage then stopped" are a failed seed fetch (SSL/
        # connection), a JS-rendered nav with no static links, or every link
        # filtered out by robots / URL rules.
        stop_reason = None
        if len(results) <= 1:
            seed_row = results[0] if results else None
            if not seed_row:
                stop_reason = 'No pages could be crawled — the start URL could not be fetched.'
            elif seed_row.get('error'):
                extra = '; '.join((seed_row.get('issues') or [])[:2])
                stop_reason = (f"Crawl stopped at the start page — it returned an error "
                               f"({seed_row.get('error')}). {extra}").strip()
            elif not (seed_row.get('internal_link_urls') or []):
                stop_reason = ("Crawl stopped at the start page — it loaded but no internal links were found. "
                               "If the site builds its navigation with JavaScript, turn on 'Render JS' and retry.")
            else:
                stop_reason = ("Crawl stopped at the start page — links were found but none were crawlable "
                               "(blocked by robots.txt, removed by your URL include/exclude filters, or pointing "
                               "to other domains). Adjust the filters or enable 'Ignore robots.txt' and retry.")

        app.logger.info(f"[crawler] Crawl complete: {len(results)} pages, {errors} errors, {avg_time}s avg, {len(dup_titles)} dup titles, {len(orphans)} orphans" + (f" | stop_reason: {stop_reason}" if stop_reason else ""))
        yield f"data: {json.dumps({'type': 'complete', 'total': len(results), 'summary': summary, 'inlinks': inlinks_payload, 'reports': reports, 'stop_reason': stop_reason})}\n\n"
        yield "data: [DONE]\n\n"
        ACTIVE_CRAWL_RULES.pop(crawl_id, None)
        ACTIVE_CRAWL_LIMITS.pop(crawl_id, None)

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Access-Control-Allow-Origin': '*'
        }
    )


# =============================================================================
# Saved crawls — store / list / load / delete / compare
# Storage: ~/.site-crawler-crawls/  (LOCAL ONLY, never pushed to git).
# Open to all users on this instance — saved crawls are shared.
# Note: this is pure file I/O + diff math. No Claude / LLM involved.
# =============================================================================

_CRAWL_FOLDER = os.path.expanduser('~/.site-crawler-crawls')
# Sibling tool's save dir — listed/loadable here too so a crawl saved
# in seo-tool shows up in site-crawler's Load saved list (and vice
# versa). Writes still go to _CRAWL_FOLDER so we don't fight over
# ownership; this is read-only union.
_CRAWL_FOLDERS_RO = [os.path.expanduser('~/.seo-tool-crawls')]

# Permanent, append-only title history. Lives inside the crawl folder as a
# .jsonl (skipped by the .json-only 30-day cleanup), so even after a full crawl
# JSON is purged, every page's title + meta description on each crawl date
# survives forever — the searchable "what was this page's title on date X"
# archive for catching title/meta regressions. One JSON line per page.
_CRAWL_TITLE_HISTORY_PATH = os.path.join(_CRAWL_FOLDER, 'crawl-titles.jsonl')


def _append_crawl_title_history(name, saved_at, results):
    """Append one line per crawled page capturing its title + meta description.
    Best-effort: never raises into the caller."""
    try:
        os.makedirs(_CRAWL_FOLDER, exist_ok=True)
        from datetime import datetime as _dt3
        ts = _dt3.utcfromtimestamp(saved_at).isoformat(timespec='seconds') + 'Z'
        seed = (results[0].get('url') if results else '') or ''
        lines = []
        for p in results:
            if not isinstance(p, dict) or not p.get('url'):
                continue
            lines.append(json.dumps({
                'ts': ts,
                'crawl': name,
                'seed': seed,
                'url': p.get('url'),
                'title': p.get('title') or '',
                'meta_description': p.get('meta_description') or '',
            }, ensure_ascii=False))
        if lines:
            with open(_CRAWL_TITLE_HISTORY_PATH, 'a') as f:
                f.write('\n'.join(lines) + '\n')
    except Exception as e:
        app.logger.warning(f'[crawl-titles] failed to append title history: {e}')


@app.route('/crawl/save', methods=['POST'])
def crawl_save():
    body = request.json or {}
    name = (body.get('name') or '').strip() or f'crawl-{int(time.time())}'
    results = body.get('results') or []
    inlinks = body.get('inlinks') or {}
    reports = body.get('reports') or {}
    if not results:
        return jsonify({'error': 'No crawl data supplied'}), 400
    os.makedirs(_CRAWL_FOLDER, exist_ok=True)
    safe = _re.sub(r'[^A-Za-z0-9._-]', '_', name)[:80]
    path = os.path.join(_CRAWL_FOLDER, f'{int(time.time())}_{safe}.json')
    try:
        with open(path, 'w') as f:
            json.dump({
                'name': name,
                'saved_at': int(time.time()),
                'pages': len(results),
                'seed': (results[0].get('url') if results else ''),
                'saved_by': (request.remote_addr or 'anon'),
                'results': results,
                'inlinks': inlinks,
                'reports': reports,
            }, f)
    except Exception as e:
        return jsonify({'error': f'Save failed: {str(e)[:200]}'}), 500
    # Permanent title history — append BEFORE the 30-day cleanup so purged
    # crawls still leave their titles on record forever.
    _append_crawl_title_history(name, int(time.time()), results)
    # 30-day cleanup
    try:
        cutoff = int(time.time()) - 30 * 86400
        for fn in os.listdir(_CRAWL_FOLDER):
            if not fn.endswith('.json'):
                continue
            fp = os.path.join(_CRAWL_FOLDER, fn)
            try:
                with open(fp) as f:
                    d = json.load(f)
            except Exception:
                continue
            if (d.get('saved_at') or 0) < cutoff:
                try: os.remove(fp)
                except OSError: pass
    except Exception:
        pass
    return jsonify({'ok': True, 'file': os.path.basename(path), 'name': name})


def _all_crawl_folders():
    """Folders to scan for /crawl/list and /crawl/load. Own dir first
    so collisions on filename prefer our own copy."""
    return [_CRAWL_FOLDER] + _CRAWL_FOLDERS_RO


def _ip_name_map():
    """Shared IP → friendly name map (lives in seo-tool's home file).
    Falls back to empty dict if the file is missing/unreadable."""
    try:
        with open(os.path.expanduser('~/.seo-tool-users.json')) as f:
            m = json.load(f)
            return m if isinstance(m, dict) else {}
    except Exception:
        return {}


@app.route('/crawl/list', methods=['GET'])
def crawl_list():
    """List ALL saved crawls from the last 30 days, unioning across this
    tool's dir + the sibling tool's dir (read-only) so users see the
    same list whichever app they open."""
    cutoff = int(time.time()) - 30 * 86400
    name_map = _ip_name_map()
    out = []
    seen_files = set()
    for folder in _all_crawl_folders():
        if not os.path.isdir(folder): continue
        source = 'seo-tool' if 'seo-tool-crawls' in folder else 'site-crawler'
        for fn in sorted(os.listdir(folder), reverse=True):
            if not fn.endswith('.json') or fn in seen_files: continue
            seen_files.add(fn)
            path = os.path.join(folder, fn)
            try:
                with open(path) as f: d = json.load(f)
                if (d.get('saved_at') or 0) < cutoff: continue
                # Cross-tool field reconciliation: site-crawler stores the
                # IP in 'saved_by'; seo-tool stores it in 'user_ip'. Read
                # both, then translate via the shared name map so the
                # column shows "Puneet" / "Yvonne" / etc instead of raw
                # 192.168.x.x (per the seo-tool-user-map convention).
                raw_ip = d.get('user_ip') or d.get('saved_by') or ''
                saved_by = name_map.get(raw_ip, '') or raw_ip or 'unknown'
                out.append({
                    'file': fn,
                    'name': d.get('name', fn),
                    'saved_at': d.get('saved_at'),
                    'pages': d.get('pages', 0),
                    'seed': d.get('seed', ''),
                    'saved_by': saved_by,
                    'source': source,
                })
            except Exception:
                continue
    out.sort(key=lambda r: r.get('saved_at') or 0, reverse=True)
    return jsonify({'crawls': out})


def _find_crawl_path(fn):
    """Locate a saved-crawl file in any of our read folders."""
    for folder in _all_crawl_folders():
        p = os.path.join(folder, fn)
        if os.path.exists(p): return p
    return None


@app.route('/crawl/load', methods=['GET'])
def crawl_load():
    fn = request.args.get('file', '')
    if not fn or '/' in fn or '\\' in fn or not fn.endswith('.json'):
        return jsonify({'error': 'Invalid file'}), 400
    path = _find_crawl_path(fn)
    if not path:
        return jsonify({'error': 'Not found'}), 404
    try:
        with open(path) as f: d = json.load(f)
    except Exception as e:
        return jsonify({'error': f'Load failed: {str(e)[:200]}'}), 500
    return jsonify({
        'results': d.get('results', []),
        'inlinks': d.get('inlinks', {}),
        'reports': d.get('reports', {}),
        'name': d.get('name', ''),
        'seed': d.get('seed', ''),
    })


@app.route('/crawl/delete', methods=['POST'])
def crawl_delete():
    fn = (request.json or {}).get('file', '')
    if not fn or '/' in fn or '\\' in fn or not fn.endswith('.json'):
        return jsonify({'error': 'Invalid file'}), 400
    path = _find_crawl_path(fn)
    if not path:
        return jsonify({'error': 'Not found'}), 404
    try:
        os.remove(path)
    except Exception as e:
        return jsonify({'error': f'Delete failed: {str(e)[:200]}'}), 500
    return jsonify({'ok': True})


@app.route('/crawl/compare', methods=['POST'])
def crawl_compare():
    """Diff two crawls. Accepts {a_file, b_file} or {a_file, b_results}.
    Returns aggregate metrics, issues comparison, structure diff,
    plus added/removed/changed URL lists."""
    body = request.json or {}

    def _load_file(fn):
        if not fn or '/' in fn or '\\' in fn or not fn.endswith('.json'):
            return None, 'Invalid file'
        fp = os.path.join(_CRAWL_FOLDER, fn)
        if not os.path.exists(fp):
            return None, 'Not found'
        try:
            with open(fp) as f: d = json.load(f)
        except Exception as e:
            return None, f'Load failed: {str(e)[:200]}'
        return d, None

    a_file = body.get('a_file', '')
    a_data, err = _load_file(a_file)
    if err: return jsonify({'error': err}), 400
    a_results = a_data.get('results', [])
    a_meta = {'name': a_data.get('name',''), 'saved_at': a_data.get('saved_at'), 'pages': a_data.get('pages', len(a_results))}

    b_file = body.get('b_file', '')
    b_results_in = body.get('b_results')
    if b_file:
        b_data, err = _load_file(b_file)
        if err: return jsonify({'error': err}), 400
        b_results = b_data.get('results', [])
        b_meta = {'name': b_data.get('name',''), 'saved_at': b_data.get('saved_at'), 'pages': b_data.get('pages', len(b_results))}
    elif isinstance(b_results_in, list):
        b_results = b_results_in
        b_meta = {'name': 'Current crawl (in memory)', 'saved_at': int(time.time()), 'pages': len(b_results)}
    else:
        return jsonify({'error': 'Supply b_file or b_results'}), 400

    def _key(u):
        return (u or '').rstrip('/').lower()
    a_by = {_key(r.get('url')): r for r in a_results if r.get('url')}
    b_by = {_key(r.get('url')): r for r in b_results if r.get('url')}
    a_urls = set(a_by.keys()); b_urls = set(b_by.keys())
    added = sorted(b_urls - a_urls); removed = sorted(a_urls - b_urls)
    shared = a_urls & b_urls

    watch = (
        'status_code', 'title', 'title_len', 'meta_description', 'meta_len',
        'h1', 'word_count', 'canonical', 'redirect_url', 'indexable',
        'depth', 'response_time', 'internal_links', 'external_links',
        'images_no_alt', 'body_hash',
    )
    def _norm_list(v):
        if v is None: return ''
        if isinstance(v, list): return ', '.join(sorted(str(x) for x in v))
        return str(v)
    changed = []
    for k in sorted(shared):
        ar = a_by[k]; br = b_by[k]
        diffs = {}
        for f in watch:
            av = ar.get(f); bv = br.get(f)
            if (av if av is not None else '') != (bv if bv is not None else ''):
                diffs[f] = {'old': av, 'new': bv}
        sa = _norm_list(ar.get('schema_types')); sb = _norm_list(br.get('schema_types'))
        if sa != sb:
            diffs['schema_types'] = {'old': sa or '—', 'new': sb or '—'}
        if diffs:
            changed.append({'url': ar.get('url') or br.get('url'), 'diffs': diffs})

    def _agg(rows):
        n = len(rows); codes = {'2xx':0,'3xx':0,'4xx':0,'5xx':0,'other':0}
        errors=warns=indexable=noindex=with_schema=redirects=missing_title=missing_meta=missing_h1=missing_canonical=title_too_long=meta_too_long=thin=slow=no_alt=0
        depths=[]; rts=[]
        for r in rows:
            sc = r.get('status_code') or 0
            if 200 <= sc < 300: codes['2xx'] += 1
            elif 300 <= sc < 400: codes['3xx'] += 1
            elif 400 <= sc < 500: codes['4xx'] += 1
            elif 500 <= sc < 600: codes['5xx'] += 1
            else: codes['other'] += 1
            if sc >= 400 or r.get('error'): errors += 1
            if r.get('issues'): warns += 1
            depths.append(r.get('depth') or 0)
            if r.get('response_time'): rts.append(r['response_time'])
            if r.get('indexable') is True: indexable += 1
            elif r.get('indexable') is False: noindex += 1
            if r.get('schema_types'): with_schema += 1
            if r.get('redirect_url'): redirects += 1
            if not r.get('title'): missing_title += 1
            elif (r.get('title_len') or 0) > 60: title_too_long += 1
            if not r.get('meta_description'): missing_meta += 1
            elif (r.get('meta_len') or 0) > 160: meta_too_long += 1
            if not r.get('h1'): missing_h1 += 1
            if not r.get('canonical'): missing_canonical += 1
            if (r.get('word_count') or 0) < 200: thin += 1
            if (r.get('response_time') or 0) > 3: slow += 1
            no_alt += int(r.get('images_no_alt') or 0)
        return {'pages':n,'codes':codes,'errors':errors,'warns_pages':warns,
                'max_depth':max(depths) if depths else 0,
                'avg_depth':round(sum(depths)/len(depths),2) if depths else 0,
                'avg_response_time':round(sum(rts)/len(rts),2) if rts else 0,
                'indexable':indexable,'noindex':noindex,'with_schema':with_schema,
                'redirects':redirects,'missing_title':missing_title,'missing_meta':missing_meta,
                'missing_h1':missing_h1,'missing_canonical':missing_canonical,
                'title_too_long':title_too_long,'meta_too_long':meta_too_long,
                'thin':thin,'slow':slow,'images_no_alt':no_alt}
    agg_a = _agg(a_results); agg_b = _agg(b_results)

    def _normalize_issue(s):
        if not s: return s
        out = _re.sub(r'\s*\([^)]*\)\s*$', '', s).strip()
        out = _re.sub(r'^\d+\s+', '', out).strip()
        return out or s
    def _issue_url_sets(rows):
        m = {}
        for r in rows:
            url = r.get('url') or ''
            seen = set()
            for issue in (r.get('issues') or []):
                norm = _normalize_issue(issue)
                if not norm or norm in seen: continue
                seen.add(norm)
                m.setdefault(norm, set()).add(url)
        return m
    urls_a_by_issue = _issue_url_sets(a_results)
    urls_b_by_issue = _issue_url_sets(b_results)
    issues_compare = []
    _URL_CAP = 500
    for iss in sorted(set(urls_a_by_issue) | set(urls_b_by_issue)):
        ua = urls_a_by_issue.get(iss, set()); ub = urls_b_by_issue.get(iss, set())
        ia = len(ua); ib = len(ub)
        if ia == ib == 0: continue
        only_a = sorted(ua - ub)[:_URL_CAP]
        only_b = sorted(ub - ua)[:_URL_CAP]
        both   = sorted(ua & ub)[:_URL_CAP]
        issues_compare.append({
            'issue': iss, 'a': ia, 'b': ib, 'delta': ib - ia,
            'only_a': only_a, 'only_b': only_b, 'both': both,
            'only_a_total': len(ua - ub),
            'only_b_total': len(ub - ua),
            'both_total':   len(ua & ub),
        })
    issues_compare.sort(key=lambda x: (abs(x['delta']), x['a'] + x['b']), reverse=True)

    from urllib.parse import urlparse as _urlp
    def _dir_counts(rows):
        c = {}
        for r in rows:
            try: p = _urlp(r.get('url') or '').path or '/'
            except Exception: p = '/'
            seg = '/' + p.lstrip('/').split('/', 1)[0] + ('/' if '/' in p.lstrip('/') else '')
            if seg in ('/', '/'): seg = '/' if p == '/' else seg
            c[seg] = c.get(seg, 0) + 1
        return c
    dirs_a = _dir_counts(a_results); dirs_b = _dir_counts(b_results)
    structure = []
    for d in sorted(set(dirs_a) | set(dirs_b)):
        da = dirs_a.get(d, 0); db = dirs_b.get(d, 0)
        if da == db == 0: continue
        structure.append({'path': d, 'a': da, 'b': db, 'delta': db - da})
    structure.sort(key=lambda x: -(x['a'] + x['b']))

    return jsonify({
        'a': a_meta, 'b': b_meta,
        'aggregate': {'a': agg_a, 'b': agg_b},
        'issues': issues_compare,
        'structure': structure[:30],
        'added': [{'url': b_by[k].get('url'), 'status_code': b_by[k].get('status_code'), 'title': b_by[k].get('title')} for k in added],
        'removed': [{'url': a_by[k].get('url'), 'status_code': a_by[k].get('status_code'), 'title': a_by[k].get('title')} for k in removed],
        'changed': changed,
        'summary': {'added': len(added), 'removed': len(removed),
                    'changed': len(changed), 'unchanged': len(shared) - len(changed)},
    })


@app.route('/crawl/update-rules', methods=['POST'])
def crawl_update_rules():
    """Apply new include/exclude patterns AND/OR per-host delay to a crawl
    that's already running.

    Merges into the existing entry — fields not present in the payload keep
    their current value. The slider can push only `crawl_delay` without
    blowing away the patterns; the patterns Apply button can push only
    include/exclude without resetting the delay.

    The crawl_id is issued in the 'start' SSE event. Patterns use robots.txt
    syntax: ``*`` is any sequence, ``$`` at end anchors end of URL, everything
    else is literal (including ``?``).
    """
    payload = request.get_json(silent=True) or {}
    crawl_id = (payload.get('crawl_id') or '').strip()
    if not crawl_id or crawl_id not in ACTIVE_CRAWL_RULES:
        return jsonify({'ok': False, 'error': 'No active crawl with that id'}), 404

    def _parse(raw):
        if not raw:
            return []
        return [p.strip() for p in raw.splitlines() if p.strip() and not p.strip().startswith('#')]

    current = ACTIVE_CRAWL_RULES.get(crawl_id) or {}
    if 'exclude_patterns' in payload:
        current['exclude'] = _parse(payload.get('exclude_patterns', ''))
    if 'include_patterns' in payload:
        current['include'] = _parse(payload.get('include_patterns', ''))
    if 'crawl_delay' in payload:
        try:
            current['crawl_delay'] = max(float(payload.get('crawl_delay')), 0.0)
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'crawl_delay must be numeric'}), 400
    ACTIVE_CRAWL_RULES[crawl_id] = current
    app.logger.info(
        f"[crawler] {crawl_id} rules updated: "
        f"{len(current.get('exclude') or [])} exclude, "
        f"{len(current.get('include') or [])} include, "
        f"delay={current.get('crawl_delay')}s"
    )
    return jsonify({
        'ok': True,
        'exclude': current.get('exclude') or [],
        'include': current.get('include') or [],
        'crawl_delay': current.get('crawl_delay'),
    })


@app.route('/crawl/resumable', methods=['GET'])
def crawl_resumable():
    """Return metadata for a suspended crawl (or 404 if unknown/expired).
    Used by the UI to decide whether to show a Resume button."""
    crawl_id = (request.args.get('crawl_id') or '').strip()
    state = SUSPENDED_CRAWLS.get(crawl_id)
    if not state:
        return jsonify({'ok': False, 'error': 'No suspended crawl with that id'}), 404
    age = int(time.time() - state.get('created', 0))
    if age > SUSPENDED_CRAWL_TTL:
        SUSPENDED_CRAWLS.pop(crawl_id, None)
        return jsonify({'ok': False, 'error': 'Suspended crawl expired'}), 410
    return jsonify({
        'ok': True,
        'crawl_id': crawl_id,
        'seed_url': state.get('seed_url'),
        'domain': state.get('domain'),
        'pages': len(state.get('results', [])),
        'queued': len(state.get('queue', [])),
        'errors': state.get('errors', 0),
        'age_seconds': age,
        'ttl_seconds': SUSPENDED_CRAWL_TTL,
    })


@app.route('/crawl/continue', methods=['POST'])
def crawl_continue():
    """Resume a crawl that's paused at the page-cap prompt.

    action=continue → bump max_pages by `bump` (default 500) and resume.
    action=finalize → stop with what we have, run summary/reports.
    Either way, signals the generator's continue_event to wake it up.
    """
    payload = request.get_json(silent=True) or {}
    crawl_id = (payload.get('crawl_id') or '').strip()
    action = (payload.get('action') or 'continue').strip().lower()
    bump = int(payload.get('bump', 500) or 500)

    state = ACTIVE_CRAWL_LIMITS.get(crawl_id)
    if not state:
        return jsonify({'ok': False, 'error': 'No paused crawl with that id'}), 404

    if action == 'finalize':
        state['finalize'] = True
    else:
        state['max_pages'] = state.get('max_pages', 0) + bump
        state['bumps'] = state.get('bumps', 0) + 1

    ev = state.get('continue_event')
    if ev is not None:
        ev.set()
    app.logger.info(f"[crawler] {crawl_id} {action}: max_pages={state.get('max_pages')}, bumps={state.get('bumps')}")
    return jsonify({'ok': True, 'max_pages': state.get('max_pages'), 'finalize': state.get('finalize'), 'bumps': state.get('bumps')})


@app.route('/export-crawl-xlsx', methods=['POST'])
def export_crawl_xlsx():
    """Export crawl results as a styled .xlsx workbook.

    Sheet 1 = All Pages (every URL, key SEO fields, color-coded status/speed).
    Sheet 2 = Issues Summary (issue text → page count → sample URLs).
    Sheet 3+ = whatever the client supplied in `extra_sheets` (built by
    `_buildExportForCategory` on the frontend so each tab gets the shape it
    actually shows). Mirrors seo-tool's export-crawl-xlsx route but without
    the Claude-specific bits.
    """
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = request.json or {}
    results = data.get('results', [])
    domain = data.get('domain', 'site')

    if not results:
        return jsonify({'error': 'No results to export'}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'All Pages'

    headers = ['URL', 'Status', 'Title', 'Title Len', 'Meta Description', 'Meta Len',
               'H1', 'Words', 'Canonical', 'Indexable', 'Int Links', 'Ext Links',
               'Images', 'Alt Missing', 'Schema', 'Speed (s)', 'Depth', 'Issues']

    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='6B5CE7', end_color='6B5CE7', fill_type='solid')
    cell_font = Font(name='Calibri', size=10)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
    )
    green_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    amber_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    red_fill   = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for ri, r in enumerate(results, 2):
        vals = [
            r.get('url', ''), r.get('status_code', 0), r.get('title', ''), r.get('title_len', 0),
            r.get('meta_description', ''), r.get('meta_len', 0), r.get('h1', ''),
            r.get('word_count', 0), r.get('canonical', ''),
            'Yes' if r.get('indexable', True) else 'No',
            r.get('internal_links', 0), r.get('external_links', 0),
            r.get('images_total', 0), r.get('images_no_alt', 0),
            ', '.join(r.get('schema_types', [])[:3]),
            r.get('response_time', 0), r.get('depth', 0),
            '; '.join(r.get('issues', []))
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = cell_font
            cell.border = thin_border
        status = r.get('status_code', 0)
        sc = ws.cell(row=ri, column=2)
        if 200 <= status < 300: sc.fill = green_fill
        elif 300 <= status < 400: sc.fill = amber_fill
        elif status >= 400: sc.fill = red_fill
        speed = r.get('response_time', 0)
        sp = ws.cell(row=ri, column=16)
        if speed <= 1: sp.fill = green_fill
        elif speed <= 3: sp.fill = amber_fill
        else: sp.fill = red_fill

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[letter].width = max_len + 3
    ws.freeze_panes = 'A2'

    # Sheet 2: Issues Summary
    ws2 = wb.create_sheet('Issues Summary')
    for ci, h in enumerate(['Issue', 'Count', 'Pages'], 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
    issue_map = {}
    for r in results:
        for issue in r.get('issues', []):
            base = issue.split('(')[0].strip()
            issue_map.setdefault(base, []).append(r.get('url', ''))
    for ri, (issue, urls) in enumerate(sorted(issue_map.items(), key=lambda x: -len(x[1])), 2):
        ws2.cell(row=ri, column=1, value=issue).font = cell_font
        ws2.cell(row=ri, column=2, value=len(urls)).font = Font(name='Calibri', size=10, bold=True)
        ws2.cell(row=ri, column=3, value='; '.join(urls[:10])).font = cell_font
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 8
    ws2.column_dimensions['C'].width = 80

    # Sheets 3+ — whatever the client built. {name, header, rows}.
    extra = data.get('extra_sheets') or []
    used_names = {ws.title, ws2.title}
    def _safe_sheet_name(n):
        s = _re.sub(r'[:\\/\?\*\[\]]', '-', str(n or 'Sheet'))[:31].strip() or 'Sheet'
        if s in used_names:
            base = s[:28]
            i = 2
            while f'{base} {i}' in used_names and i < 99:
                i += 1
            s = f'{base} {i}'
        used_names.add(s)
        return s
    for sheet in extra:
        name = _safe_sheet_name(sheet.get('name') or 'Sheet')
        header = sheet.get('header') or []
        rows = sheet.get('rows') or []
        if not header and not rows:
            continue
        ws_e = wb.create_sheet(name)
        for ci, h in enumerate(header, 1):
            cell = ws_e.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        for ri, r in enumerate(rows, 2):
            for ci, v in enumerate(r, 1):
                if isinstance(v, (list, tuple)):
                    v = ', '.join(str(x) for x in v)
                elif isinstance(v, dict):
                    v = json.dumps(v, ensure_ascii=False)
                cell = ws_e.cell(row=ri, column=ci, value=v)
                cell.font = cell_font
                cell.border = thin_border
        for col in ws_e.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), 80))
            ws_e.column_dimensions[letter].width = max(8, max_len + 2)
        ws_e.freeze_panes = 'A2'

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_domain = _re.sub(r'[^a-zA-Z0-9.-]', '', (domain or '').lstrip('.').replace('www.', '', 1))
    from datetime import datetime as _dt
    _ts = _dt.now().strftime('%Y-%m-%d-%H%M')
    _name = '-'.join([p for p in ['crawl', safe_domain, _ts] if p]) + '.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=_name,
    )


@app.route('/export-crawl-sitemap', methods=['POST'])
def export_crawl_sitemap():
    """Build a sitemaps.org 0.9 XML sitemap from the crawl results.

    Filter: include only URLs that a search engine would actually want to
    index - HTTP 200, indexable (no noindex meta or X-Robots-Tag), self- or
    missing-canonical (skip pages that canonical to a different URL since
    those aren't the canonical version), and not redirected. >50K URLs
    splits into a sitemap-index plus N child sitemaps to stay within the
    sitemaps.org 50,000-URL / 50MB-uncompressed limits.
    """
    from xml.sax.saxutils import escape as _xml_escape
    from datetime import datetime as _dt
    from email.utils import parsedate_to_datetime as _parsedate
    import io as _io
    import zipfile as _zipfile

    data = request.json or {}
    results = data.get('results', [])
    domain = (data.get('domain') or '').strip()

    if not results:
        return json.dumps({'error': 'No results to export'}), 400

    def _is_indexable_200(r):
        if r.get('status_code') != 200:
            return False
        if r.get('error'):
            return False
        if r.get('redirect_url'):
            return False
        if r.get('indexable') is False:
            return False
        if r.get('canonical_kind') == 'canonicalised':
            return False
        ctype = (r.get('content_type') or '').lower()
        if ctype and 'html' not in ctype and 'xml' not in ctype:
            return False
        if r.get('is_pagination'):
            return False
        url = (r.get('url') or '').strip()
        if not url or not (url.startswith('http://') or url.startswith('https://')):
            return False
        return True

    eligible = [r for r in results if _is_indexable_200(r)]

    seen = set()
    urls = []
    for r in eligible:
        u = r['url'].strip()
        if u in seen:
            continue
        seen.add(u)
        # Format Last-Modified as W3C date (YYYY-MM-DD). Header arrives as
        # RFC 1123; fall back silently if missing or malformed.
        lastmod = ''
        lm_raw = (r.get('last_modified') or '').strip()
        if lm_raw:
            try:
                lastmod = _parsedate(lm_raw).strftime('%Y-%m-%d')
            except Exception:
                lastmod = ''
        urls.append({'loc': u, 'lastmod': lastmod})

    SITEMAP_NS = 'http://www.sitemaps.org/schemas/sitemap/0.9'
    URLS_PER_SITEMAP = 50000

    def _build_urlset(chunk):
        lines = ['<?xml version="1.0" encoding="UTF-8"?>',
                 f'<urlset xmlns="{SITEMAP_NS}">']
        for u in chunk:
            lines.append('  <url>')
            lines.append(f'    <loc>{_xml_escape(u["loc"])}</loc>')
            if u['lastmod']:
                lines.append(f'    <lastmod>{u["lastmod"]}</lastmod>')
            lines.append('  </url>')
        lines.append('</urlset>')
        lines.append('')
        return '\n'.join(lines)

    safe_domain = _re.sub(r'[^a-zA-Z0-9.-]', '', (domain or '').lstrip('.').replace('www.', '', 1))
    _ts = _dt.now().strftime('%Y-%m-%d-%H%M')

    if len(urls) <= URLS_PER_SITEMAP:
        xml = _build_urlset(urls)
        _name = '-'.join([p for p in ['sitemap', safe_domain, _ts] if p]) + '.xml'
        return send_file(_io.BytesIO(xml.encode('utf-8')),
            mimetype='application/xml',
            as_attachment=True,
            download_name=_name)

    zip_buf = _io.BytesIO()
    today = _dt.now().strftime('%Y-%m-%d')
    with _zipfile.ZipFile(zip_buf, 'w', _zipfile.ZIP_DEFLATED) as zf:
        index_lines = ['<?xml version="1.0" encoding="UTF-8"?>',
                       f'<sitemapindex xmlns="{SITEMAP_NS}">']
        for i in range(0, len(urls), URLS_PER_SITEMAP):
            chunk = urls[i:i + URLS_PER_SITEMAP]
            child_name = f'sitemap-{i // URLS_PER_SITEMAP + 1}.xml'
            zf.writestr(child_name, _build_urlset(chunk))
            host = domain.rstrip('/') if domain else ''
            index_lines.append('  <sitemap>')
            index_lines.append(f'    <loc>{_xml_escape(host + "/" + child_name)}</loc>')
            index_lines.append(f'    <lastmod>{today}</lastmod>')
            index_lines.append('  </sitemap>')
        index_lines.append('</sitemapindex>')
        index_lines.append('')
        zf.writestr('sitemap.xml', '\n'.join(index_lines))
    zip_buf.seek(0)
    _name = '-'.join([p for p in ['sitemap', safe_domain, _ts] if p]) + '.zip'
    return send_file(zip_buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=_name)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5002, debug=False, threaded=True)

